"""
GKIN Amstelveen Mededelingen Generator – Web App
Redeploy trigger
"""

import base64
import builtins
import os
import shutil
import tempfile
import threading
import traceback
import uuid
from datetime import datetime, timedelta
from typing import Dict, Optional
from flask import Flask, render_template, request, send_file, jsonify, session, redirect, url_for
from werkzeug.utils import secure_filename

import re
import requests as _requests
from bs4 import BeautifulSoup as _BS
import pandas as pd
import dropbox

from data_sources.dropbox_reader import DropboxExcelReader
from data_sources.scipio_scraper import ScipioScraper
from data_sources.preekroster_scraper import PreekrosterScraper
from data_sources.email_reader import OutlookCollecteReader, get_token_cache_json
from bulletin_generator import BulletinGenerator
from voorlees_generator import VoorleesGenerator

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB max upload
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'gkin-amstelveen-secret-2026')
SITE_PASSWORD = os.environ.get('SITE_PASSWORD', '')

def _password_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if SITE_PASSWORD and not session.get('authenticated'):
            return redirect(url_for('login_page', next=request.path))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    error = None
    if request.method == 'POST':
        if request.form.get('password') == SITE_PASSWORD:
            session['authenticated'] = True
            next_url = request.args.get('next') or '/'
            return redirect(next_url)
        error = 'Onjuist wachtwoord. Probeer opnieuw.'
    return render_template('login.html', error=error,
                           next=request.args.get('next', '/'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'output', '_uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Cache takenrooster on startup
_takenrooster_cache = None

# Cache bijbelbasics KND data (slug -> {title, verse, date_day, date_month})
_knd_cache = None
_KND_DUTCH_MONTHS = {
    'januari':1,'februari':2,'maart':3,'april':4,'mei':5,'juni':6,
    'juli':7,'augustus':8,'september':9,'oktober':10,'november':11,'december':12
}


def _load_knd_cache():
    global _knd_cache
    if _knd_cache is not None:
        return _knd_cache
    try:
        r = _requests.get('https://www.bijbelbasics.nl/programma/', timeout=10)
        soup = _BS(r.text, 'html.parser')
        lines = [l.strip() for l in soup.get_text(separator='\n').split('\n') if l.strip()]
        date_pat = re.compile(
            r'^(?:zondag|zaterdag|vrijdag)\s+(\d+)\s+(\w+)\s+(\d{4})$', re.I)
        verse_start = re.compile(r'^([A-Z][a-zA-Zë]+)\s+(\d+:\d+)$')
        verse_end   = re.compile(r'^-\s*(\d+:\d+)$')
        entries = []
        for i, line in enumerate(lines):
            m = date_pat.match(line)
            if not m:
                continue
            day       = int(m.group(1))
            month_num = _KND_DUTCH_MONTHS.get(m.group(2).lower(), 0)
            year      = int(m.group(3))
            if not month_num:
                continue
            # title is the line before the date
            title = lines[i - 1] if i > 0 else ''
            # verse: next two lines are "Book X:Y" and "- X:Z"
            verse = ''
            if i + 2 < len(lines):
                ms = verse_start.match(lines[i + 1])
                me = verse_end.match(lines[i + 2])
                if ms and me:
                    verse = f"{ms.group(1)} {ms.group(2)} - {me.group(1)}"
                elif ms:
                    verse = f"{ms.group(1)} {ms.group(2)}"
            entries.append({'title': title, 'verse': verse,
                            'day': day, 'month': month_num, 'year': year})
        _knd_cache = entries
    except Exception as e:
        # Don't cache on error - allow retry on next request
        print(f'[KND] Error loading from bijbelbasics.nl: {e}')
        return []
    return _knd_cache


def _get_takenrooster():
    global _takenrooster_cache
    if _takenrooster_cache is None:
        reader = DropboxExcelReader()
        _takenrooster_cache = reader.get_takenrooster()
    return _takenrooster_cache


def _get_takenrooster_and_render_page():
    taken = _get_takenrooster()
    dutch_months = [
        'januari', 'februari', 'maart', 'april', 'mei', 'juni',
        'juli', 'augustus', 'september', 'oktober', 'november', 'december'
    ]
    dutch_days = ['maandag', 'dinsdag', 'woensdag', 'donderdag',
                  'vrijdag', 'zaterdag', 'zondag']

    today = datetime.now().date()
    dates = []
    for entry in taken['entries']:
        d = entry['date']
        if hasattr(d, 'date'):
            d_date = d.date()
        else:
            d_date = d
        if d_date < today:
            continue
        opm_med = entry.get('opmerking', '')
        opm_clean_med = ''
        if opm_med:
            opm_clean_med = opm_med.split('OLE')[0].strip().rstrip(',').strip()
            if not opm_clean_med and 'OLE' in opm_med:
                opm_clean_med = 'OLE'
        suffix = f"  —  {entry['predikant']}"
        if opm_clean_med:
            suffix += f" ({opm_clean_med})"
        label = f"{dutch_days[d.weekday()]} {d.day} {dutch_months[d.month - 1]} {d.year}{suffix}"
        dates.append({
            'value': d.strftime('%Y-%m-%d'),
            'label': label,
            'day_idx':   d.weekday(),
            'month_idx': d.month - 1,
            'day_num':   d.day,
            'year':      d.year,
            'suffix':    suffix,
            'predikant': entry['predikant'],
            'ovd': entry.get('ovd', ''),
            'predikant_email': entry.get('predikant_email', ''),
            'beamer_email':    entry.get('beamer_email', ''),
            'beamer':          entry.get('beamer', ''),
            '1eo':             entry.get('1eo', ''),
            '1eo_email':       entry.get('1eo_email', ''),
        })

    return render_template('mededelingen.html', dates=dates)


@app.route('/')
def home():
    return render_template('home.html')


@app.route('/mededelingen')
@_password_required
def mededelingen_index():
    return _get_takenrooster_and_render_page()


@app.route('/preekbevestiging')
@_password_required
def preekbevestiging_index():
    taken = _get_takenrooster()
    dutch_months = [
        'januari', 'februari', 'maart', 'april', 'mei', 'juni',
        'juli', 'augustus', 'september', 'oktober', 'november', 'december'
    ]
    dutch_days = ['maandag', 'dinsdag', 'woensdag', 'donderdag',
                  'vrijdag', 'zaterdag', 'zondag']
    today = datetime.now().date()
    dates = []
    for entry in taken['entries']:
        d = entry['date']
        d_date = d.date() if hasattr(d, 'date') else d
        if d_date < today:
            continue
        opm = entry.get('opmerking', '')
        opm_clean = ''
        if opm:
            opm_clean = opm.split('OLE')[0].strip().rstrip(',').strip()
            if not opm_clean and 'OLE' in opm:
                opm_clean = 'OLE'
        pb_suffix = f"  —  {entry['predikant']}"
        if opm_clean:
            pb_suffix += f" ({opm_clean})"
        label = f"{dutch_days[d.weekday()]} {d.day} {dutch_months[d.month - 1]} {d.year}{pb_suffix}"
        dates.append({
            'value':        d.strftime('%Y-%m-%d'),
            'label':        label,
            'day_idx':      d.weekday(),
            'month_idx':    d.month - 1,
            'day_num':      d.day,
            'year':         d.year,
            'suffix':       pb_suffix,
            'predikant':    entry['predikant'],
            'ovd':          entry.get('ovd', ''),
            'ovd_email':    entry.get('ovd_email', ''),
            '1eo':          entry.get('1eo', ''),
            '1eo_email':    entry.get('1eo_email', ''),
            'beamer':       entry.get('beamer', ''),
            'beamer_email': entry.get('beamer_email', ''),
            'dag':          entry.get('dag', ''),
            'tijd':            entry.get('tijd', '10:30') or '10:30',
            'opmerking':       opm,
            'predikant_email': entry.get('predikant_email', ''),
            'muziek':          entry.get('muziek', ''),
            'voorzangers':  entry.get('voorzangers', ''),
            'multimedia':   entry.get('multimedia', ''),
            'knd':          entry.get('knd', ''),
            'tieners':      entry.get('tieners', ''),
            'unresolved_names': entry.get('unresolved_names', []),
        })
    return render_template('preekbevestiging.html', dates=dates)


@app.route('/generate-preekbevestiging', methods=['POST'])
def generate_preekbevestiging():
    from preekbevestiging_generator import generate as gen_doc
    data     = request.get_json(force=True)
    iso_date = data.get('date', '')
    if not iso_date:
        return jsonify({'error': 'no date'}), 400
    taken = _get_takenrooster()
    entry = next((e for e in taken['entries']
                  if e['date'].strftime('%Y-%m-%d') == iso_date), None)
    if not entry:
        return jsonify({'error': 'date not found'}), 404
    d = datetime.strptime(iso_date, '%Y-%m-%d')
    filename = f"Preekbevestiging_{d.strftime('%Y%m%d')}_{entry.get('predikant','').replace(' ','_')}.docx"
    out_path = os.path.join(UPLOAD_DIR, filename)
    songs = data.get('songs', [])
    gen_doc(entry, iso_date, out_path, songs=songs)
    return send_file(out_path, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')


@app.route('/fetch-liederen', methods=['POST'])
def fetch_liederen():
    from data_sources.email_reader import OutlookCollecteReader
    data     = request.get_json(force=True)
    iso_date = data.get('date', '')
    if not iso_date:
        return jsonify({'error': 'no date'}), 400
    try:
        d = datetime.strptime(iso_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date'}), 400
    try:
        reader = OutlookCollecteReader()
        if not reader.is_authenticated():
            return jsonify({'error': 'not_authenticated', 'songs': ['','','','','','',''], 'not_found': ['Niet ingelogd bij Outlook']}), 200
        result = reader.fetch_liederen(target_date=d)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e), 'songs': ['','','','','','',''], 'not_found': [str(e)]}), 200


@app.route('/fetch-language', methods=['POST'])
def fetch_language():
    from data_sources.preekroster_scraper import PreekrosterScraper
    data     = request.get_json(force=True)
    iso_date = data.get('date', '')
    if not iso_date:
        return jsonify({'error': 'no date'}), 400
    try:
        d = datetime.strptime(iso_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date'}), 400

    dutch_months_short = ['jan','feb','mrt','apr','mei','jun',
                          'jul','aug','sep','okt','nov','dec']
    date_key = f"{d.day} {dutch_months_short[d.month - 1]}"  # e.g. "17 mei"

    try:
        from datetime import timedelta
        scraper = PreekrosterScraper()
        # Use d - 6 days as mededelingen_date so d always falls inside the window
        roster  = scraper.get_preekroster(mededelingen_date=d - timedelta(days=6))
        for entry in roster.get('am_table', []):
            raw = entry.get('date', '')
            # raw may be "28 jun" or "28 jun (Pinksteren)"
            if raw.startswith(date_key):
                return jsonify({'language': entry.get('language', '')})
    except Exception:
        pass
    return jsonify({'language': ''})


@app.route('/fetch-ole-preekroster', methods=['POST'])
def fetch_ole_preekroster():
    """Fetch OLE preekroster data for a given date - includes location, time, predikant."""
    from data_sources.preekroster_scraper import PreekrosterScraper
    data = request.get_json(force=True)
    iso_date = data.get('date', '')
    if not iso_date:
        return jsonify({'error': 'no date'}), 400
    try:
        d = datetime.strptime(iso_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date'}), 400

    dutch_months_short = ['jan','feb','mrt','apr','mei','jun',
                          'jul','aug','sep','okt','nov','dec']
    date_key = f"{d.day} {dutch_months_short[d.month - 1]}"  # e.g. "17 mei"

    try:
        from datetime import timedelta
        scraper = PreekrosterScraper()
        roster = scraper.get_preekroster(mededelingen_date=d - timedelta(days=6))
        
        # Look for OLE entry matching this date
        for entry in roster.get('ole_table', []):
            raw = entry.get('date', '')
            if raw.startswith(date_key):
                return jsonify({
                    'found': True,
                    'predikant': entry.get('predikant', ''),
                    'location': entry.get('regio', ''),  # AM, DH, TB, etc.
                    'time': entry.get('time', '10.00u'),
                    'language': entry.get('language', ''),
                    'full_date': entry.get('date', '')
                })
        
        return jsonify({
            'found': False,
            'error': 'Geen OLE dienst gevonden voor deze datum in preekroster'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/fetch-knd-thema', methods=['POST'])
def fetch_knd_thema():
    data = request.get_json(force=True)
    date_str = data.get('date', '')
    if not date_str:
        return jsonify({'error': 'no date'}), 400
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date'}), 400
    entries = _load_knd_cache()
    for entry in entries:
        if entry['day'] == d.day and entry['month'] == d.month:
            return jsonify({'title': entry['title'], 'verse': entry['verse']})
    return jsonify({'title': '', 'verse': ''})


@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    """Serve files from the uploads directory (for QR image previews)."""
    return send_file(os.path.join(UPLOAD_DIR, filename))


@app.route('/generate', methods=['POST'])
def generate():
    date_str = request.form.get('date')
    if not date_str:
        return jsonify({'error': 'No date selected'}), 400

    selected_date = datetime.strptime(date_str, '%Y-%m-%d')

    try:
        # 1. Get takenrooster entry
        taken = _get_takenrooster()
        entry = None
        for e in taken['entries']:
            if e['date'] == selected_date:
                entry = e
                break
        if entry is None:
            return jsonify({'error': f'No takenrooster entry for {date_str}'}), 404

        # 2. Mededelingen from Dropbox
        reader = DropboxExcelReader()
        meded = reader.get_mededelingen(mededelingen_date=selected_date)

        # 3. Birthdays from Scipio
        scipio = ScipioScraper()
        bdays = scipio.get_birthday_list(mededelingen_date=selected_date)

        # 4. Preekroster from GKIN website
        preek = PreekrosterScraper()
        roster = preek.get_preekroster(mededelingen_date=selected_date)

        # 5. Collect user-supplied manual fields from form
        def fv(key): return request.form.get(key, '').strip()

        # Parse activiteiten rows (datum_0, activiteit_0, locatie_0, ...)
        activiteiten = []
        i = 0
        while request.form.get(f'act_datum_{i}') is not None:
            row = {
                'datum': fv(f'act_datum_{i}'),
                'tijd': fv(f'act_tijd_{i}'),
                'activiteit': fv(f'act_activiteit_{i}'),
                'olv': fv(f'act_olv_{i}'),
                'locatie': fv(f'act_locatie_{i}'),
            }
            if any(row.values()):
                activiteiten.append(row)
            i += 1

        # Handle QR image uploads — prefer file upload, fall back to email-fetched path
        def save_upload(field_name, fallback_field):
            f = request.files.get(field_name)
            if f and f.filename:
                ext = os.path.splitext(secure_filename(f.filename))[1].lower()
                path = os.path.join(UPLOAD_DIR, f'{uuid.uuid4().hex}{ext}')
                f.save(path)
                return path
            # Use email-fetched filename if present
            fname = fv(fallback_field)
            if fname:
                full = os.path.join(UPLOAD_DIR, os.path.basename(fname))
                if os.path.exists(full):
                    return full
            return ''

        user_data = {
            'overdenking_predikant':      fv('overdenking_predikant'),
            'overdenking_thema':          fv('overdenking_thema'),
            'overdenking_schriftlezing':  fv('overdenking_schriftlezing'),
            'overdenking_content':        fv('overdenking_content'),
            'collecte_contant':           fv('collecte_contant'),
            'collecte_bonnen':            fv('collecte_bonnen'),
            'collecte_bank':              fv('collecte_bank'),
            'collecte_tikkie':            fv('collecte_tikkie'),
            'collecte_ole':               fv('collecte_ole'),
            'bezoekers_volwassenen':      fv('bezoekers_volwassenen'),
            'bezoekers_kinderen':         fv('bezoekers_kinderen'),
            'dankoffer_url':              fv('dankoffer_url'),
            'dankoffer_qr':               save_upload('dankoffer_qr_file', 'dankoffer_qr_path'),
            'ole_url':                    fv('ole_url'),
            'ole_qr':                     save_upload('ole_qr_file', 'ole_qr_path'),
            'activiteiten':               activiteiten,
            'opbrengst_entries':          [],  # filled below if email entries were fetched
        }

        # Re-assemble multi-entry opbrengst from hidden JSON field (set by JS after email fetch)
        import json as _json
        opbrengst_json = request.form.get('opbrengst_entries_json', '')
        if opbrengst_json:
            try:
                user_data['opbrengst_entries'] = _json.loads(opbrengst_json)
            except Exception:
                pass

        # 6. Generate bulletin
        gen = BulletinGenerator()
        filepath = gen.generate(selected_date, entry, taken['entries'],
                                meded, bdays, roster, user_data)

        filename = os.path.basename(filepath)
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/generate-voorlees', methods=['POST'])
def generate_voorlees():
    date_str = request.form.get('date')
    if not date_str:
        return jsonify({'error': 'No date selected'}), 400
    selected_date = datetime.strptime(date_str, '%Y-%m-%d')
    try:
        taken = _get_takenrooster()
        entry = None
        for e in taken['entries']:
            if e['date'] == selected_date:
                entry = e
                break
        if entry is None:
            return jsonify({'error': f'Geen rooster gevonden voor {date_str}'}), 404

        reader = DropboxExcelReader()
        meded  = reader.get_mededelingen(mededelingen_date=selected_date)

        def fv(key): return request.form.get(key, '').strip()

        def save_upload_v(field_name, fallback_field):
            f = request.files.get(field_name)
            if f and f.filename:
                ext  = os.path.splitext(secure_filename(f.filename))[1].lower()
                path = os.path.join(UPLOAD_DIR, f'{uuid.uuid4().hex}{ext}')
                f.save(path)
                return path
            fname = fv(fallback_field)
            if fname:
                full = os.path.join(UPLOAD_DIR, os.path.basename(fname))
                if os.path.exists(full):
                    return full
            return ''

        import json as _json
        opbrengst_entries = []
        opbrengst_json = request.form.get('opbrengst_entries_json', '')
        if opbrengst_json:
            try:
                opbrengst_entries = _json.loads(opbrengst_json)
            except Exception:
                pass

        user_data = {
            'collecte_contant':      fv('collecte_contant'),
            'collecte_bonnen':       fv('collecte_bonnen'),
            'collecte_bank':         fv('collecte_bank'),
            'collecte_tikkie':       fv('collecte_tikkie'),
            'collecte_ole':          fv('collecte_ole'),
            'bezoekers_volwassenen': fv('bezoekers_volwassenen'),
            'bezoekers_kinderen':    fv('bezoekers_kinderen'),
            'dankoffer_url':         fv('dankoffer_url'),
            'dankoffer_qr':          save_upload_v('dankoffer_qr_file', 'dankoffer_qr_path'),
            'ole_url':               fv('ole_url'),
            'ole_qr':                save_upload_v('ole_qr_file', 'ole_qr_path'),
            'opbrengst_entries':     opbrengst_entries,
        }

        # Extract welkomstwoord paragraphs from the mededelingen template
        welkom_paras = _extract_welkom_paragraphs(selected_date, entry, meded)

        gen      = VoorleesGenerator()
        filepath = gen.generate(selected_date, entry, meded, user_data,
                                welkom_paragraphs=welkom_paras)
        filename = os.path.basename(filepath)
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _extract_welkom_paragraphs(selected_date: datetime, entry: dict, meded: dict) -> list:
    """Read the Welkomstwoord paragraphs from the mededelingen Word template."""
    from bulletin_generator import BulletinGenerator, TEMPLATE_PATH
    from docx import Document as _Document
    try:
        doc = _Document(TEMPLATE_PATH)
        paras = []
        in_welkom = False
        for p in doc.paragraphs:
            if p.style.name.startswith('Heading') and 'welkomstwoord' in p.text.lower():
                in_welkom = True
                continue
            if in_welkom:
                if p.style.name.startswith('Heading'):
                    break
                txt = p.text.strip()
                if txt:
                    paras.append(txt)
        # Fill in dynamic date/predikant/ovd from the template text
        predikant = entry.get('predikant', '')
        ovd = entry.get('ovd', '')
        dutch_months = ['januari','februari','maart','april','mei','juni',
                        'juli','augustus','september','oktober','november','december']
        dutch_days = ['maandag','dinsdag','woensdag','donderdag','vrijdag','zaterdag','zondag']
        date_str = f"{selected_date.day} {dutch_months[selected_date.month-1]} {selected_date.year}"
        day_name = dutch_days[selected_date.weekday()]
        result = []
        for p in paras:
            if 'Vandaag,' in p:
                p = f"Vandaag, {day_name} {date_str}, gaat voor {predikant}. De ouderling van dienst is {ovd}. Als u vragen heeft, kunt u de ouderling van dienst aanspreken."
            result.append(p)
        return result
    except Exception:
        return []


@app.route('/get-mededelingen', methods=['POST'])
def get_mededelingen_data():
    """Fetch mededelingen for a date; load activities from the year tab of the Excel file."""
    date_str = request.form.get('date')
    if not date_str:
        return jsonify({'error': 'No date'}), 400
    selected_date = datetime.strptime(date_str, '%Y-%m-%d')
    try:
        reader = DropboxExcelReader()
        meded = reader.get_mededelingen(mededelingen_date=selected_date)
        activities = reader.get_activiteiten_kalender(mededelingen_date=selected_date)
        return jsonify({
            'regionale_nl': meded.get('regionale_nl', ''),
            'landelijke_nl': meded.get('landelijke_nl', ''),
            'activities': activities,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _parse_activities_from_mededelingen(meded: dict, selected_date: datetime = None) -> list:
    """Parse activity rows from regionale + landelijke mededelingen blocks.

    Scans each paragraph block for Dutch date mentions (e.g. '21 juni', 'zondag 24 mei 2026').
    Each block that contains a date becomes one activity row.
    """
    import re

    MONTHS = r'(?:januari|februari|maart|april|mei|juni|juli|augustus|september|oktober|november|december|jan|feb|mrt|apr|jun|jul|aug|sep|okt|nov|dec)'
    DAYS   = r'(?:maandag|dinsdag|woensdag|donderdag|vrijdag|zaterdag|zondag|ma|di|wo|do|vr|za|zo)'

    # Matches: optional day-name, day-number, month-name, optional year
    date_pat = re.compile(
        rf'(?:{DAYS}\s+)?(\d{{1,2}}\s+{MONTHS}\w*(?:\s+\d{{4}})?)',
        re.IGNORECASE
    )
    # Matches time like "10.30 uur" or "10:30 uur"
    time_pat = re.compile(r'\b(\d{1,2}[:.]\d{2})\s*uur\b', re.IGNORECASE)
    # Matches "o.l.v. Name", "onder leiding van Name", "geleid door Name", "waarin Name voorgaat"
    # Stop at: " op [weekday/date]", " voorgaat", " spreekt", " zal", comma, newline
    olv_pat  = re.compile(
        r'(?:o\.?l\.?v\.?|onder leiding van|geleid door|waarbij\s+|waarin\s+)'
        r'(.*?)'
        r'(?=\s+(?:voorgaat|spreekt|zal\b|op\s+(?:zondag|maandag|dinsdag|woensdag|donderdag|vrijdag|zaterdag|\d))|[,\n]|$)',
        re.IGNORECASE
    )
    # Location: "in VenueName Number" or "te City" — require uppercase start to avoid "in rouw", "bijwonen"
    loc_pat  = re.compile(
        r'(?:\bin\s+([A-Z][A-Za-zÀ-ÿ\s]+\s+\d+)'  # "in Bouwerij 52" — uppercase start
        r'|\bte\s+([A-Z][A-Za-zÀ-ÿ\s]{2,20}?)(?=[,.\n ]|$))'  # "te Den Haag"
    )

    # Keywords that indicate a block is NOT a scheduled activity
    SKIP_KEYWORDS = re.compile(
        r'\b(overlijden|uitvaart|geboren|condoleance|in rouw|bijzondere gelegenheid|leden in)\b',
        re.IGNORECASE
    )

    activities = []
    sources = [meded.get('regionale_nl', ''), meded.get('landelijke_nl', '')]

    for source in sources:
        blocks = [b.strip() for b in source.split('\n\n') if b.strip()]
        for block in blocks:
            # Skip obituaries and non-activity blocks
            if SKIP_KEYWORDS.search(block):
                continue

            flat = ' '.join(block.split('\n'))

            dm = date_pat.search(flat)
            if not dm:
                continue

            datum = re.sub(r'\s+\d{4}$', '', dm.group(1).strip())
            # Shorten month names to 3-letter abbreviations: "juni" → "jun", "maart" → "mrt"
            MONTH_ABBR = {'januari':'jan','februari':'feb','maart':'mrt','april':'apr',
                          'juni':'jun','juli':'jul','augustus':'aug','september':'sep',
                          'oktober':'okt','november':'nov','december':'dec'}
            for full, abbr in MONTH_ABBR.items():
                datum = re.sub(rf'\b{full}\b', abbr, datum, flags=re.IGNORECASE)

            # Filter out dates in the past relative to selected_date
            if selected_date:
                MONTH_NUM = {'jan':1,'feb':2,'mrt':3,'mar':3,'apr':4,'mei':5,'jun':6,
                             'jul':7,'aug':8,'sep':9,'okt':10,'nov':11,'dec':12}
                dm2 = re.match(r'(\d{1,2})\s+(\w+)', datum)
                if dm2:
                    try:
                        day = int(dm2.group(1))
                        mon = MONTH_NUM.get(dm2.group(2).lower()[:3], 0)
                        yr  = selected_date.year
                        act_date = datetime(yr, mon, day) if mon else None
                        if act_date and act_date.date() < selected_date.date():
                            continue
                    except Exception:
                        pass

            # Time
            tm = time_pat.search(flat)
            tijd = tm.group(1).replace('.', ':') if tm else ''

            # o.l.v.
            ovm = olv_pat.search(flat)
            if ovm:
                raw_olv = ovm.group(1).strip()
                # Replace abbreviation dots temporarily (titles + single/double initials)
                raw_olv = re.sub(r'\b(ds|br|zr|dr|drs|mr|ir|prof)\.\s*', r'\1___', raw_olv, flags=re.IGNORECASE)
                raw_olv = re.sub(r'\b([A-Z])\.\s*', r'\1___', raw_olv)  # initials like S.M.
                raw_olv = re.split(r',|\.\s+', raw_olv)[0]
                raw_olv = raw_olv.replace('___', '. ').strip().rstrip('–- ').strip()
                olv = raw_olv
            else:
                olv = ''

            # Activity name: first line heading
            first_line = block.split('\n')[0].strip().rstrip(':–-').strip()
            activiteit = first_line

            # Location: prefer "in StreetName Number" then "te City"
            locatie = ''
            for lm in loc_pat.finditer(flat):
                cand = (lm.group(1) or lm.group(2) or '').strip().rstrip('.,')
                # Only keep short, clean strings
                if cand and len(cand) <= 30:
                    locatie = cand
                    break

            activities.append({
                'datum': datum,
                'tijd': tijd,
                'activiteit': activiteit,
                'olv': olv,
                'locatie': locatie,
            })

    return activities


@app.route('/fetch-email-collecte', methods=['POST'])
def fetch_email_collecte():
    """Fetch collecte URLs, QR images and amounts from recent Outlook emails."""
    try:
        reader = OutlookCollecteReader()
        if not reader.is_authenticated():
            return jsonify({
                'error': 'login_required',
                'message': 'Nog niet ingelogd bij Outlook. Klik op "Inloggen bij Outlook" eerst.'
            }), 401
        # Parse selected date from request
        date_str = request.json.get('date', '') if request.is_json else ''
        target_date = None
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                pass
        data = reader.fetch_collecte_data(target_date=target_date, since_days=60)
        return jsonify(data)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/fetch-email-overdenking', methods=['POST'])
def fetch_email_overdenking():
    """Fetch overdenking content from scriba email attachment."""
    try:
        reader = OutlookCollecteReader()
        if not reader.is_authenticated():
            return jsonify({'error': 'login_required'}), 401
        date_str = request.json.get('date', '') if request.is_json else ''
        target_date = None
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                pass
        data = reader.fetch_overdenking(target_date=target_date, since_days=14)
        return jsonify(data)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/fetch-email-opbrengst', methods=['POST'])
def fetch_email_opbrengst():
    """Fetch collecte opbrengsten amounts and bezoekers from Outlook emails."""
    try:
        reader = OutlookCollecteReader()
        if not reader.is_authenticated():
            return jsonify({'error': 'login_required'}), 401
        date_str = request.json.get('date', '') if request.is_json else ''
        target_date = None
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                pass
        data = reader.fetch_opbrengst_data(target_date=target_date, since_days=60)
        return jsonify(data)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/email-login-start', methods=['POST'])
def email_login_start():
    """Start device-code OAuth2 flow; returns user_code and verification_uri."""
    try:
        reader = OutlookCollecteReader()
        flow   = reader.start_device_flow()
        return jsonify(flow)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/email-login-complete', methods=['POST'])
def email_login_complete():
    """Poll for completed device-code login."""
    try:
        reader  = OutlookCollecteReader()
        success = reader.complete_device_flow()
        if success:
            return jsonify({'status': 'ok'})
        return jsonify({'status': 'pending'}), 202
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/email-auth-status', methods=['GET'])
def email_auth_status():
    """Check whether a valid Outlook token is already cached."""
    try:
        reader = OutlookCollecteReader()
        return jsonify({'authenticated': reader.is_authenticated()})
    except Exception as e:
        return jsonify({'authenticated': False, 'error': str(e)})


@app.route('/email-token-export', methods=['GET'])
def email_token_export():
    """Export current MSAL token cache as JSON string."""
    try:
        token_json = get_token_cache_json()
        if not token_json or token_json == '{}':
            return jsonify({'error': 'Geen token gevonden — log eerst in bij Outlook.'}), 404
        return jsonify({'token_cache': token_json,
                        'instructions': 'Gebruik /email-token-download voor een schoon tekstbestand.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/email-token-download', methods=['GET'])
def email_token_download():
    """Download the raw MSAL token cache as a plain .txt file.
    Open the file and paste its entire contents into Railway > Variables > MSAL_TOKEN_CACHE.
    """
    try:
        from flask import Response
        token_json = get_token_cache_json()
        if not token_json or token_json == '{}':
            return jsonify({'error': 'Geen token gevonden — log eerst in bij Outlook.'}), 404
        return Response(
            token_json,
            mimetype='text/plain',
            headers={'Content-Disposition': 'attachment; filename=msal_token_cache.txt'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/refresh-dates', methods=['POST'])
def refresh_dates():
    """Force refresh of takenrooster cache."""
    global _takenrooster_cache
    _takenrooster_cache = None
    return jsonify({'status': 'ok'})


# ---------------------------------------------------------------------------
# Liturgie routes
# ---------------------------------------------------------------------------

_LITURGIE_BASE    = os.path.dirname(__file__)
_LITURGIE_CACHE   = os.path.join(_LITURGIE_BASE, 'bible_cache')
_LITURGIE_LOGO    = os.path.join(_LITURGIE_BASE, 'logo.png')
_LITURGIE_PHONE   = os.path.join(_LITURGIE_BASE, 'telephone.gif')
_LITURGIE_LOCK    = threading.Lock()
_LITURGIE_FILE_LOCKS: dict = {}

DROPBOX_APP_KEY_L      = os.getenv('DROPBOX_APP_KEY', '')
DROPBOX_APP_SECRET_L   = os.getenv('DROPBOX_APP_SECRET', '')
DROPBOX_REFRESH_L      = os.getenv('DROPBOX_REFRESH_TOKEN', '')
BIBLE_DROPBOX_PATH     = '/working folder/bible'
LOGO_DROPBOX_PATH_L    = '/working folder/logo.png'
PHONE_DROPBOX_PATH_L   = '/working folder/telephone.gif'


def _get_dbx_liturgie():
    import dropbox
    return dropbox.Dropbox(
        oauth2_refresh_token=DROPBOX_REFRESH_L,
        app_key=DROPBOX_APP_KEY_L,
        app_secret=DROPBOX_APP_SECRET_L,
    )


def _ensure_liturgie_file(remote_path: str, local_path: str):
    """Download a single file from Dropbox if not cached locally."""
    if os.path.exists(local_path):
        return
    try:
        dbx = _get_dbx_liturgie()
        _, resp = dbx.files_download(remote_path)
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, 'wb') as f:
            f.write(resp.content)
    except Exception as e:
        print(f'Warning: could not download {remote_path}: {e}')


def _ensure_bible_file(filename: str):
    """Download a single bible chapter file from Dropbox on demand."""
    os.makedirs(_LITURGIE_CACHE, exist_ok=True)
    local_path = os.path.join(_LITURGIE_CACHE, filename)
    if os.path.exists(local_path):
        return
    with _LITURGIE_LOCK:
        if filename not in _LITURGIE_FILE_LOCKS:
            _LITURGIE_FILE_LOCKS[filename] = threading.Lock()
    with _LITURGIE_FILE_LOCKS[filename]:
        if os.path.exists(local_path):
            return
        dbx = _get_dbx_liturgie()
        dropbox_path = f'{BIBLE_DROPBOX_PATH}/{filename}'
        try:
            _, resp = dbx.files_download(dropbox_path)
            with open(local_path, 'wb') as f:
                f.write(resp.content)
        except Exception as e:
            raise FileNotFoundError(f'Bijbelbestand niet gevonden in Dropbox: {filename} ({e})')


def _run_liturgi(excel_bytes: bytes, preek_bytes, work_dir: str) -> dict:
    _ensure_liturgie_file(LOGO_DROPBOX_PATH_L, _LITURGIE_LOGO)
    _ensure_liturgie_file(PHONE_DROPBOX_PATH_L, _LITURGIE_PHONE)
    os.makedirs(_LITURGIE_CACHE, exist_ok=True)

    file_mingguan = os.path.join(work_dir, 'file mingguan')
    os.makedirs(file_mingguan, exist_ok=True)

    with open(os.path.join(file_mingguan, 'Main Liturgy file.xlsx'), 'wb') as f:
        f.write(excel_bytes)
    if preek_bytes:
        with open(os.path.join(file_mingguan, 'Preek.docx'), 'wb') as f:
            f.write(preek_bytes)

    for name, src in [('bible', _LITURGIE_CACHE), ('logo.png', _LITURGIE_LOGO), ('telephone.gif', _LITURGIE_PHONE)]:
        link = os.path.join(work_dir, name)
        if not os.path.exists(link) and os.path.exists(src):
            os.symlink(src, link)

    src_path = os.path.join(_LITURGIE_BASE, 'liturgi_core.py')
    with open(src_path, 'r', encoding='utf-8') as fh:
        patched = fh.read().replace(
            'dir_path = os.path.dirname(os.path.realpath(__file__))',
            f'dir_path = {repr(work_dir)}'
        )

    ns = {
        '__file__': src_path,
        '__name__': '__liturgi_run__',
        '__builtins__': builtins,
        '_ensure_bible_file': _ensure_bible_file,
    }
    try:
        exec(compile(patched, src_path, 'exec'), ns)
    except SystemExit:
        pass

    result = {}
    for fname in os.listdir(file_mingguan):
        if fname.startswith('LiturgieA') and fname.endswith('.docx'):
            result['liturgieA'] = os.path.join(file_mingguan, fname)
            result['liturgieA_name'] = fname
        elif fname.startswith('LiturgieB') and fname.endswith('.docx'):
            result['liturgieB'] = os.path.join(file_mingguan, fname)
            result['liturgieB_name'] = fname
        elif fname.startswith('LiturgieP') and fname.endswith('.pptx'):
            result['liturgieP'] = os.path.join(file_mingguan, fname)
            result['liturgieP_name'] = fname
    return result


@app.route('/liturgie')
def liturgie_index():
    return render_template('liturgie.html')


PREEK_DROPBOX_PATH = '/working folder/file mingguan/Preek.docx'

@app.route('/liturgie/generate', methods=['POST'])
def liturgie_generate():
    excel_source = request.form.get('excel_source', 'upload')
    preek_source = request.form.get('preek_source', 'upload')
    want_a = request.form.get('want_a', '1') == '1'
    want_b = request.form.get('want_b', '1') == '1'
    want_p = request.form.get('want_p', '1') == '1'

    # Get Excel bytes
    if excel_source == 'dropbox':
        try:
            dbx = _get_dbx_liturgie()
            _, resp = dbx.files_download(WORKING_FILE_PATH)
            excel_bytes = resp.content
        except Exception as e:
            return jsonify({'error': f'Kon Main Liturgy file niet laden van Dropbox: {e}'}), 500
    else:
        excel_file = request.files.get('excel_file')
        if not excel_file or not excel_file.filename:
            return jsonify({'error': 'Excel bestand is verplicht.'}), 400
        excel_bytes = excel_file.read()

    # Get Preek bytes
    if preek_source == 'dropbox':
        try:
            dbx = _get_dbx_liturgie()
            _, resp = dbx.files_download(PREEK_DROPBOX_PATH)
            preek_bytes = resp.content
        except Exception as e:
            preek_bytes = None  # Preek is optional, continue without it
            print(f'[Generate] Could not load Preek.docx from Dropbox: {e}')
    else:
        preek_file = request.files.get('preek_file')
        preek_bytes = preek_file.read() if preek_file and preek_file.filename else None

    work_dir = tempfile.mkdtemp(prefix='liturgi_')
    try:
        result = _run_liturgi(excel_bytes, preek_bytes, work_dir)
        if not result:
            return jsonify({'error': 'Geen output bestanden gegenereerd.'}), 500

        files_out = []
        for key, want in [('liturgieA', want_a), ('liturgieB', want_b), ('liturgieP', want_p)]:
            if want and key in result:
                with open(result[key], 'rb') as fh:
                    files_out.append({'name': result[f'{key}_name'], 'data': base64.b64encode(fh.read()).decode('ascii')})

        if not files_out:
            return jsonify({'error': 'Geen geselecteerde bestanden gevonden in output.'}), 500
        return jsonify(files_out)

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Liturgie Auto-fill Endpoint
# ---------------------------------------------------------------------------

DANKOFFER_DROPBOX_PATH = '/working folder/Dankoffer.xlsx'

def _get_dankoffer_verse(dbx, service_date: datetime, mark_as_used: bool = True) -> Optional[Dict]:
    """
    Read Dankoffer.xlsx from Dropbox and return the next available verse.

    Expected Dankoffer.xlsx format:
    - Column A: Bible verse reference (e.g., "Psalmen 50:14-15")
    - Column B: Bible text (the actual verse content)
    - Column C: Date Used (blank = not used yet, or shows date when used)

    Logic:
    1. Find the first row where Column C (Date Used) is blank/empty
    2. Return that verse
    3. If mark_as_used=True, update Column C with the service date
    4. If all verses are used, reset all dates and use the first verse

    Returns dict with: book, chapter, verse_start, verse_end, full_text, row_index, total_count
    """
    from io import BytesIO

    try:
        _, resp = dbx.files_download(DANKOFFER_DROPBOX_PATH)
        df = pd.read_excel(BytesIO(resp.content), header=None)

        # Check if first row is a header
        first_cell = str(df.iloc[0, 0]).strip().lower() if pd.notna(df.iloc[0, 0]) else ''
        header_keywords = ['verse', 'bible', 'text', 'reference', 'ref', 'book', 'chapter', 'date', 'gebruikt']
        has_header = any(keyword in first_cell for keyword in header_keywords)

        if has_header:
            # Skip header row
            df = df.iloc[1:].reset_index(drop=True)
            print(f'[Dankoffer] Header detected. Data starts at row 1 (skipping header)')

        # Read verses and their used dates from Column C (index 2)
        verses_data = []
        for idx, row in df.iterrows():
            verse = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
            # Column C (index 2) contains Date Used
            date_used = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ''

            if verse and verse.lower() not in ('nan', 'none', ''):
                verses_data.append({
                    'verse': verse,
                    'date_used': date_used if date_used.lower() not in ('nan', 'none', '') else '',
                    'row_idx': idx
                })

        if not verses_data:
            return None

        # Format the service date for comparison (YYYY-MM-DD)
        service_date_str = service_date.strftime('%Y-%m-%d')
        print(f'[Dankoffer] Looking for service date: {service_date_str}')
        print(f'[Dankoffer] Verses data: {len(verses_data)} verses loaded')

        # STEP 1: Check if this service date is already assigned to a verse
        existing_assignment = None
        for v in verses_data:
            print(f'[Dankoffer] Checking verse row {v["row_idx"]}: date_used="{v["date_used"]}" vs service_date="{service_date_str}"')
            if v['date_used'] == service_date_str:
                existing_assignment = v
                print(f'[Dankoffer] ✓ Found existing assignment at row {v["row_idx"]}')
                break

        if existing_assignment:
            # This date already has a verse assigned - use it again
            selected = existing_assignment
            reset_needed = False
            already_assigned = True
            print(f'[Dankoffer] Reusing existing verse: {selected["verse"]}')
        else:
            # STEP 2: Find the first unused verse (blank date)
            already_assigned = False
            unused_verses = [v for v in verses_data if not v['date_used']]
            print(f'[Dankoffer] No existing assignment found. {len(unused_verses)} unused verses available')

            if unused_verses:
                # Use the first unused verse
                selected = unused_verses[0]
                reset_needed = False
            else:
                # STEP 3: All verses are used - find the one with the OLDEST date
                # Parse dates and find the minimum (oldest)
                dated_verses = []
                for v in verses_data:
                    try:
                        if v['date_used']:
                            parsed_date = datetime.strptime(v['date_used'], '%Y-%m-%d')
                            dated_verses.append({**v, 'parsed_date': parsed_date})
                    except ValueError:
                        # If date format is invalid, treat as very old
                        dated_verses.append({**v, 'parsed_date': datetime(1900, 1, 1)})

                if dated_verses:
                    # Sort by date and pick the oldest
                    dated_verses.sort(key=lambda x: x['parsed_date'])
                    selected = dated_verses[0]
                    reset_needed = True  # Indicates we're reusing an old verse
                else:
                    # Fallback - should not happen
                    selected = verses_data[0]
                    reset_needed = True

        verse_text = selected['verse']

        # Parse verse text like "Psalmen 50:14-15" or "Psalmen 50:14"
        import re
        match = re.match(r'^(.+?)\s+(\d+):(\d+)(?:-(\d+))?$', verse_text.strip())
        if not match:
            return None

        book = match.group(1).strip()
        chapter = match.group(2)
        verse_start = match.group(3)
        verse_end = match.group(4)  # None if single verse

        # Calculate unused count (for display purposes)
        if already_assigned:
            unused_count = len([v for v in verses_data if not v['date_used']])
        else:
            unused_count = len(unused_verses) if not reset_needed else 0

        result = {
            'book': book,
            'chapter': chapter,
            'verse_start': verse_start,
            'verse_end': verse_end,
            'full_text': verse_text,
            'row_index': selected['row_idx'] + 1,  # 1-indexed for user display
            'total_count': len(verses_data),
            'unused_count': unused_count,
            'reset_needed': reset_needed,
            'already_assigned': already_assigned,
            'date_assigned': selected['date_used'] if already_assigned else None
        }

        # Update the Dankoffer.xlsx to mark this verse as used (or update date if reusing)
        if mark_as_used:
            try:
                _mark_dankoffer_verse_as_used(dbx, selected['row_idx'], service_date, reset_needed)
                result['marked_as_used'] = True
            except Exception as e:
                print(f'[Dankoffer] Warning: Could not mark verse as used: {e}')
                result['marked_as_used'] = False

        return result

    except Exception as e:
        print(f'[Dankoffer] Error reading dankoffer file: {e}')
        return None


def _mark_dankoffer_verse_as_used(dbx, row_idx: int, service_date: datetime, reset_all: bool = False):
    """
    Update Dankoffer.xlsx in Dropbox to mark a verse as used (Column C).
    Updates just the selected row with the new date.
    Auto-detects if there's a header row based on first cell content.
    """
    try:
        from io import BytesIO
        import openpyxl

        # Download current file
        _, resp = dbx.files_download(DANKOFFER_DROPBOX_PATH)
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active

        # Detect if row 1 is a header by checking if A1 looks like a header
        first_cell = str(ws.cell(row=1, column=1).value).strip().lower() if ws.cell(row=1, column=1).value else ''
        # If A1 contains words like "verse", "bible", "text", "reference", it's likely a header
        header_keywords = ['verse', 'bible', 'text', 'reference', 'ref', 'book', 'chapter', 'date', 'gebruikt']
        has_header = any(keyword in first_cell for keyword in header_keywords)

        # Calculate Excel row: if header exists, data starts at row 2
        if has_header:
            excel_row = row_idx + 2  # row_idx 0 -> Excel row 2 (first data row after header)
            print(f'[Dankoffer] Header detected. Writing to row {excel_row} (row_idx={row_idx})')
        else:
            excel_row = row_idx + 1  # row_idx 0 -> Excel row 1
            print(f'[Dankoffer] No header detected. Writing to row {excel_row} (row_idx={row_idx})')

        date_str = service_date.strftime('%Y-%m-%d')

        # Read current value for logging
        current_val = ws.cell(row=excel_row, column=3).value
        print(f'[Dankoffer] Current value in row {excel_row}, col 3: {current_val}')

        # Update the cell
        ws.cell(row=excel_row, column=3).value = date_str
        print(f'[Dankoffer] Setting row {excel_row}, col 3 to: {date_str}')

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Upload back to Dropbox
        dbx.files_upload(output.read(), DANKOFFER_DROPBOX_PATH, mode=dropbox.files.WriteMode('overwrite'))
        print(f'[Dankoffer] Successfully marked verse at row {excel_row}, column C as used on {date_str}')

    except Exception as e:
        print(f'[Dankoffer] Error updating dankoffer file: {e}')
        import traceback
        traceback.print_exc()
        raise


def _preview_dankoffer_verse(dbx, service_date: datetime) -> Optional[Dict]:
    """Preview which verse will be used without marking it as used."""
    return _get_dankoffer_verse(dbx, service_date, mark_as_used=False)


def _parse_verse_reference(verse_text: str) -> Optional[Dict]:
    """Parse a verse reference like 'Psalmen 50:14-15' or 'Psalmen 50:14'."""
    import re
    match = re.match(r'^(.+?)\s+(\d+):(\d+)(?:-(\d+))?$', verse_text.strip())
    if not match:
        return None

    return {
        'book': match.group(1).strip(),
        'chapter': match.group(2),
        'verse_start': match.group(3),
        'verse_end': match.group(4),  # None if single verse
    }


@app.route('/liturgie/fill-data', methods=['POST'])
def liturgie_fill_data():
    """
    Auto-populate Main Liturgy file.xlsx with:
    - People on duty from Takenrooster (B4-B12)
    - Dankoffer verse from Dankoffer.xlsx (B21-E21)
    - Tikkie link from email (if available)

    Uses openpyxl to preserve formatting, images, and other content.
    Only modifies specific cells without changing the rest of the file.

    Returns a new Excel file with populated data and alert info.
    """
    excel_file = request.files.get('excel_file')
    if not excel_file or not excel_file.filename:
        return jsonify({'error': 'Excel bestand is verplicht.'}), 400

    excel_bytes = excel_file.read()

    try:
        # Use openpyxl to preserve formatting and images
        from io import BytesIO
        import openpyxl

        # Load workbook preserving all features including data validation
        wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=False, keep_links=True)
        ws = wb['Data'] if 'Data' in wb.sheetnames else wb.active

        # Get date from B3 (row 3, column 2 in openpyxl 1-indexed)
        date_cell = ws.cell(row=3, column=2)
        date_val = date_cell.value

        if not date_val:
            return jsonify({'error': 'Geen datum gevonden in cel B3.'}), 400

        # Parse the date
        if isinstance(date_val, datetime):
            service_date = date_val
        else:
            try:
                service_date = datetime.strptime(str(date_val), '%d-%m-%Y') if '-' in str(date_val) else datetime.strptime(str(date_val), '%Y-%m-%d')
            except ValueError:
                try:
                    service_date = pd.to_datetime(str(date_val), dayfirst=True).to_pydatetime()
                except Exception:
                    return jsonify({'error': f'Ongeldige datum in cel B3: {date_val}'}), 400

        # Get takenrooster entry for this date
        taken = _get_takenrooster()
        entry = None
        for e in taken.get('entries', []):
            entry_date = e['date']
            if hasattr(entry_date, 'date'):
                entry_date = entry_date.date()
            else:
                entry_date = entry_date
            if entry_date == service_date.date():
                entry = e
                break

        if not entry:
            return jsonify({
                'error': f'Geen dienst gevonden in takenrooster voor {service_date.strftime("%d-%m-%Y")}'
            }), 404

        # Track what was already filled vs what we populated
        alerts = {
            'already_filled': [],
            'auto_populated': [],
            'not_found': []
        }

        # Helper function to get cell value as string
        def get_cell_value(row, col):
            val = ws.cell(row=row, column=col).value
            return str(val).strip() if val else ''

        # Helper function to set cell value while preserving formatting
        def set_cell_value(row, col, value, field_name):
            cell = ws.cell(row=row, column=col)
            current_val = str(cell.value).strip() if cell.value else ''

            if current_val and current_val.lower() not in ('nan', 'none', ''):
                alerts['already_filled'].append(f'{field_name}: {current_val}')
                return False
            elif value:
                cell.value = value
                alerts['auto_populated'].append(f'{field_name}: {value}')
                return True
            return False

        # Define the fields to populate (excel row, field name, takenrooster key, excel column)
        # B4-B12 correspond to rows 4-12, column 2
        field_mapping = [
            (4, 'Voorganger', 'predikant', 2),      # B4
            (5, 'OvD', 'ovd', 2),                   # B5
            (6, '1e Ontvangst', '1eo', 2),          # B6
            (7, 'Muzikanten', 'muziek', 2),         # B7
            (8, 'Voorzangers', 'voorzangers', 2),   # B8
            (9, 'Beamer', 'beamer', 2),             # B9
            (10, 'Geluid', 'multimedia', 2),        # B10
            (11, 'KND', 'knd', 2),                  # B11
            (12, 'Tieners', 'tieners', 2),          # B12
        ]

        # Populate B4-B12
        for row, field_name, taken_key, col in field_mapping:
            new_val = str(entry.get(taken_key, '')).strip()
            if not new_val:
                alerts['not_found'].append(f'{field_name}: niet gevonden in takenrooster')
            elif new_val:
                set_cell_value(row, col, new_val, field_name)

        # Get dankoffer verse from Dropbox
        dbx = _get_dbx_liturgie()
        dankoffer = _get_dankoffer_verse(dbx, service_date)

        # Row 21 (dankoffer) - B21-E21
        dankoffer_row = 21
        print(f'[Liturgie Fill] Dankoffer data: {dankoffer}')

        if dankoffer:
            # Read Boeken sheet for validation (search all columns)
            valid_books = []
            if 'Boeken' in wb.sheetnames:
                boeken_ws = wb['Boeken']
                for row in range(1, boeken_ws.max_row + 1):
                    # Check all columns (1-3) for book names
                    for col in range(1, 4):
                        book_name = boeken_ws.cell(row=row, column=col).value
                        if book_name and isinstance(book_name, str):
                            book_str = book_name.strip()
                            if book_str and len(book_str) > 1:
                                valid_books.append(book_str)
                # Remove duplicates while preserving order
                seen = set()
                unique_books = []
                for book in valid_books:
                    book_lower = book.lower()
                    if book_lower not in seen:
                        seen.add(book_lower)
                        unique_books.append(book)
                valid_books = unique_books
                print(f'[Liturgie Fill] Found {len(valid_books)} valid books in Boeken sheet (all columns)')
                print(f'[Liturgie Fill] First 15 books: {valid_books[:15]}')

            # Validate dankoffer book name against Boeken list
            dankoffer_book = dankoffer['book']
            book_valid = False
            if valid_books:
                # Check exact match or partial match (case-insensitive)
                dankoffer_book_normalized = dankoffer_book.lower().replace(' ', '').replace('.', '').replace('ii', '2')
                print(f'[Liturgie Fill] Looking for book: "{dankoffer_book}" (normalized: "{dankoffer_book_normalized}")')

                for valid_book in valid_books:
                    valid_book_normalized = valid_book.lower().replace(' ', '').replace('.', '').replace('ii', '2')

                    # Check various match types
                    exact_match = dankoffer_book.lower() == valid_book.lower()
                    contains_match = dankoffer_book.lower() in valid_book.lower() or valid_book.lower() in dankoffer_book.lower()
                    normalized_match = dankoffer_book_normalized == valid_book_normalized

                    if exact_match or contains_match or normalized_match:
                        book_valid = True
                        # Use the exact name from Boeken list for consistency
                        dankoffer_book = valid_book
                        print(f'[Liturgie Fill] ✓ Validated book name: "{dankoffer_book}" (match type: exact={exact_match}, contains={contains_match}, normalized={normalized_match})')
                        break

                if not book_valid:
                    alerts['already_filled'].append(f'⚠️ Dankoffer boek "{dankoffer["book"]}" NIET gevonden in Boeken lijst! Controleer spelling.')
                    print(f'[Liturgie Fill] WARNING: Book "{dankoffer["book"]}" not found in {len(valid_books)} books')
                    print(f'[Liturgie Fill] Boeken list sample: {valid_books[40:50] if len(valid_books) > 50 else valid_books}')

            # Check current values in dankoffer cells
            current_b21 = get_cell_value(dankoffer_row, 2)
            current_c21 = get_cell_value(dankoffer_row, 3)
            current_d21 = get_cell_value(dankoffer_row, 4)
            current_e21 = get_cell_value(dankoffer_row, 5)
            print(f'[Liturgie Fill] Current dankoffer values - B21: "{current_b21}", C21: "{current_c21}", D21: "{current_d21}", E21: "{current_e21}"')

            # Set all dankoffer cells (directly, without individual alerts)
            dankoffer_filled_parts = []
            
            # Helper to set cell without adding to alerts
            def set_dankoffer_cell(row, col, value):
                cell = ws.cell(row=row, column=col)
                current_val = str(cell.value).strip() if cell.value else ''
                if current_val and current_val.lower() not in ('nan', 'none', ''):
                    return False  # Already has content
                elif value:
                    cell.value = value
                    return True
                return False
            
            # B21: Book name (only if validation passed or no Boeken sheet)
            if book_valid or not valid_books:
                result_b21 = set_dankoffer_cell(dankoffer_row, 2, dankoffer_book)
                print(f'[Liturgie Fill] B21 set result: {result_b21}, value: {dankoffer_book}')
                if result_b21:
                    dankoffer_filled_parts.append(dankoffer_book)
            else:
                result_b21 = False
                print(f'[Liturgie Fill] B21 NOT set - book validation failed')

            # C21: Chapter (H.S. / pasal)
            result_c21 = set_dankoffer_cell(dankoffer_row, 3, dankoffer['chapter'])
            print(f'[Liturgie Fill] C21 set result: {result_c21}, value: {dankoffer["chapter"]}')
            if result_c21:
                dankoffer_filled_parts.append(dankoffer['chapter'])

            # D21: Start verse (ayat)
            result_d21 = set_dankoffer_cell(dankoffer_row, 4, dankoffer['verse_start'])
            print(f'[Liturgie Fill] D21 set result: {result_d21}, value: {dankoffer["verse_start"]}')
            if result_d21:
                verse_text = dankoffer['verse_start']
                if dankoffer['verse_end']:
                    verse_text += f'-{dankoffer["verse_end"]}'
                dankoffer_filled_parts.append(verse_text)

            # E21: End verse (ayat) - only if there's an end verse
            if dankoffer['verse_end']:
                result_e21 = set_dankoffer_cell(dankoffer_row, 5, dankoffer['verse_end'])
                print(f'[Liturgie Fill] E21 set result: {result_e21}, value: {dankoffer["verse_end"]}')

            # Add simplified one-line dankoffer alert
            if dankoffer_filled_parts:
                alerts['auto_populated'].append(f'Dankoffer vers: {dankoffer["full_text"]}')

            # Store results for response
            dankoffer_cells = {
                'B21_filled': result_b21,
                'C21_filled': result_c21,
                'D21_filled': result_d21,
                'E21_filled': result_e21 if dankoffer['verse_end'] else False,
                'B21_value': dankoffer['book'],
                'C21_value': dankoffer['chapter'],
                'D21_value': dankoffer['verse_start'],
                'E21_value': dankoffer['verse_end'] if dankoffer['verse_end'] else None
            }
        else:
            print(f'[Liturgie Fill] No dankoffer data returned!')
            dankoffer_cells = None

        # Try to get Tikkie link from email if available
        tikkie_found = False
        try:
            reader = OutlookCollecteReader()
            if reader.is_authenticated():
                email_data = reader.fetch_collecte_data(target_date=service_date, since_days=60)
                tikkie_url = email_data.get('dankoffer_url', '')
                if tikkie_url:
                    tikkie_found = True
                    # Find Tikkie link row by looking for "Tikkie link" label in column A
                    tikkie_row = None
                    for row in range(1, ws.max_row + 1):
                        label = str(ws.cell(row=row, column=1).value).strip().lower() if ws.cell(row=row, column=1).value else ''
                        if 'tikkie' in label or 'qr_link' in label:
                            tikkie_row = row
                            break

                    if tikkie_row is not None:
                        set_cell_value(tikkie_row, 2, tikkie_url, 'Tikkie link')
                else:
                    alerts['not_found'].append('Tikkie link niet gevonden in e-mail')
            else:
                alerts['not_found'].append('Outlook niet geauthenticeerd - Tikkie link niet opgehaald')
        except Exception as e:
            print(f'[Liturgie Fill] Could not fetch Tikkie link: {e}')
            alerts['not_found'].append(f'Tikkie link ophalen mislukt: {str(e)}')

        # Save the modified Excel preserving all formatting
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Encode for response
        excel_b64 = base64.b64encode(output.read()).decode('ascii')

        # Prepare dankoffer info for response
        dankoffer_info = None
        if dankoffer:
            # Build status detail message for blue box
            status_parts = []
            if dankoffer.get('already_assigned'):
                status_parts.append('Dit vers was al toegewezen aan deze datum')
            elif dankoffer.get('marked_as_used'):
                status_parts.append('Nieuwe toewijzing toegevoegd aan Dankoffer.xlsx')
            
            if dankoffer.get('reset_needed'):
                status_parts.append('Alle verzen waren gebruikt → oudste verzen worden hergebruikt')
            
            status_parts.append(f'Rij {dankoffer["row_index"]} van {dankoffer["total_count"]} ({dankoffer["unused_count"]} resterend)')
            
            dankoffer_info = {
                'verse': dankoffer['full_text'],
                'row_index': dankoffer['row_index'],
                'total_count': dankoffer['total_count'],
                'unused_count': dankoffer['unused_count'],
                'reset_needed': dankoffer.get('reset_needed', False),
                'marked_as_used': dankoffer.get('marked_as_used', False),
                'already_assigned': dankoffer.get('already_assigned', False),
                'date_assigned': dankoffer.get('date_assigned'),
                'status_detail': ' • '.join(status_parts),
                'cells': dankoffer_cells
            }

        return jsonify({
            'excel_data': excel_b64,
            'filename': 'Main Liturgy file.xlsx',
            'alerts': alerts,
            'service_date': service_date.strftime('%d-%m-%Y'),
            'dankoffer_verse': dankoffer['full_text'] if dankoffer else None,
            'dankoffer_info': dankoffer_info
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Dankoffer Preview Endpoint
# ---------------------------------------------------------------------------

@app.route('/liturgie/preview-dankoffer', methods=['POST'])
def preview_dankoffer():
    """
    Preview which dankoffer verse will be used for a given date.
    Does NOT mark the verse as used - just shows what would be selected.

    Request body: { "date": "2026-05-11" } (ISO format)
    """
    data = request.get_json(force=True)
    date_str = data.get('date', '')

    if not date_str:
        return jsonify({'error': 'Geen datum opgegeven.'}), 400

    try:
        service_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Ongeldige datum formaat. Gebruik YYYY-MM-DD.'}), 400

    try:
        dbx = _get_dbx_liturgie()
        dankoffer = _preview_dankoffer_verse(dbx, service_date)

        if not dankoffer:
            return jsonify({
                'error': 'Geen dankoffer verzen gevonden in Dankoffer.xlsx.'
            }), 404

        return jsonify({
            'verse': dankoffer['full_text'],
            'row_index': dankoffer['row_index'],
            'total_count': dankoffer['total_count'],
            'unused_count': dankoffer['unused_count'],
            'reset_needed': dankoffer.get('reset_needed', False),
            'service_date': service_date.strftime('%d-%m-%Y'),
            'message': f'Vers {dankoffer["row_index"]} van {dankoffer["total_count"]} zal gebruikt worden.' +
                       (' (ALLE verzen zijn al gebruikt - cyclus wordt gereset!)' if dankoffer.get('reset_needed') else '')
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Liturgie Auto-Fill Preview Endpoint
# ---------------------------------------------------------------------------

@app.route('/liturgie/preview-fill-data', methods=['POST'])
def preview_liturgie_fill_data():
    """
    Preview what changes would be made to Main Liturgy file.xlsx
    Reads from working folder and returns preview without saving.
    """
    try:
        # Read Excel from working folder
        working_file_path = '/working/folder/file mingguan/Main Liturgy file.xlsx'
        
        # For now, since we can't access the working folder directly in this context,
        # we'll need the user to upload the file first, then we'll cache it
        # In production, this would read directly from the working folder
        
        if 'excel_data' not in request.files:
            return jsonify({'error': 'Geen Excel bestand geüpload. Upload eerst Main Liturgy file.xlsx'}), 400
            
        excel_file = request.files['excel_data']
        excel_bytes = excel_file.read()
        
        # Load with openpyxl to preserve formatting
        from openpyxl import load_workbook
        from io import BytesIO
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb['Data'] if 'Data' in wb.sheetnames else wb.active
        
        # Get service date from Data sheet B3
        service_date_raw = ws.cell(row=3, column=2).value
        if not service_date_raw:
            return jsonify({'error': 'Geen datum gevonden in cel B3 (Data tab)'}), 400
            
        # Parse date
        if isinstance(service_date_raw, str):
            try:
                service_date = datetime.strptime(service_date_raw, '%d-%m-%Y')
            except ValueError:
                try:
                    service_date = datetime.strptime(service_date_raw, '%Y-%m-%d')
                except ValueError:
                    return jsonify({'error': f'Ongeldige datum formaat in B3: {service_date_raw}'}), 400
        else:
            service_date = service_date_raw
            
        # Preview data - get from sources without saving
        preview = {
            'service_date': service_date.strftime('%d-%m-%Y'),
            'current_values': {},
            'proposed_changes': {},
            'alerts': {
                'not_found': [],
                'warnings': []
            }
        }
        
        # Get current values from Excel
        # B4-B12: People on duty
        field_mapping = [
            (4, 'Voorganger'),       # B4
            (5, 'OvD'),              # B5
            (6, '1e Ontvangst'),     # B6
            (7, '2e Ontvangst'),     # B7
            (8, 'Voorzangers'),      # B8
            (9, 'Beamer'),           # B9
            (10, 'Geluid'),          # B10
            (11, 'KND'),             # B11
            (12, 'Tieners'),         # B12
        ]
        
        for row, field_name in field_mapping:
            current_val = str(ws.cell(row=row, column=2).value).strip() if ws.cell(row=row, column=2).value else ''
            preview['current_values'][field_name] = current_val
            
        # Current Tikkie link
        tikkie_row = None
        for row in range(1, ws.max_row + 1):
            label = str(ws.cell(row=row, column=1).value).strip().lower() if ws.cell(row=row, column=1).value else ''
            if 'tikkie' in label or 'qr_link' in label:
                tikkie_row = row
                break
        if tikkie_row:
            current_tikkie = str(ws.cell(row=tikkie_row, column=2).value).strip() if ws.cell(row=tikkie_row, column=2).value else ''
            preview['current_values']['Tikkie link'] = current_tikkie
            
        # Current Dankoffer verse
        dankoffer_book = str(ws.cell(row=21, column=2).value).strip() if ws.cell(row=21, column=2).value else ''
        dankoffer_chapter = str(ws.cell(row=21, column=3).value).strip() if ws.cell(row=21, column=3).value else ''
        dankoffer_verse = str(ws.cell(row=21, column=4).value).strip() if ws.cell(row=21, column=4).value else ''
        if dankoffer_book and dankoffer_chapter and dankoffer_verse:
            preview['current_values']['Dankoffer vers'] = f'{dankoffer_book} {dankoffer_chapter}:{dankoffer_verse}'
        else:
            preview['current_values']['Dankoffer vers'] = ''
        
        # Get proposed values from sources
        try:
            takenrooster_data = _get_takenrooster()
            takenrooster = takenrooster_data.get('entries', [])
            entry = None
            for tr_entry in takenrooster:
                entry_date = tr_entry.get('date')
                if entry_date and entry_date.date() == service_date.date():
                    entry = tr_entry
                    break
            
            if entry:
                taken_mapping = {
                    'Voorganger': 'predikant',
                    'OvD': 'ovd',
                    '1e Ontvangst': '1eo',
                    '2e Ontvangst': '2eo',
                    'Voorzangers': 'voorzangers',
                    'Beamer': 'beamer',
                    'Geluid': 'multimedia',
                    'KND': 'knd',
                    'Tieners': 'tieners',
                }
                
                for field_name, key in taken_mapping.items():
                    new_val = str(entry.get(key, '')).strip()
                    if new_val:
                        preview['proposed_changes'][field_name] = new_val
            else:
                preview['alerts']['not_found'].append(f'Geen dienst gevonden in takenrooster voor {service_date.strftime("%d-%m-%Y")}')
        except Exception as e:
            preview['alerts']['warnings'].append(f'Fout bij ophalen takenrooster: {str(e)}')
        
        # Get Tikkie link
        try:
            reader = OutlookCollecteReader()
            if reader.is_authenticated():
                email_data = reader.fetch_collecte_data(target_date=service_date, since_days=60)
                tikkie_url = email_data.get('dankoffer_url', '')
                if tikkie_url:
                    preview['proposed_changes']['Tikkie link'] = tikkie_url
                else:
                    preview['alerts']['not_found'].append('Tikkie link niet gevonden in e-mail')
            else:
                preview['alerts']['warnings'].append('Outlook niet geauthenticeerd - Tikkie link niet opgehaald')
        except Exception as e:
            preview['alerts']['warnings'].append(f'Fout bij ophalen Tikkie: {str(e)}')
        
        # Get Dankoffer verse
        try:
            dbx = _get_dbx_liturgie()
            dankoffer = _get_dankoffer_verse(dbx, service_date, mark_as_used=False)  # Preview only, don't mark
            
            if dankoffer:
                verse_text = f"{dankoffer['book']} {dankoffer['chapter']}:{dankoffer['verse_start']}"
                if dankoffer['verse_end']:
                    verse_text += f"-{dankoffer['verse_end']}"
                preview['proposed_changes']['Dankoffer vers'] = verse_text
                preview['dankoffer_info'] = {
                    'verse': verse_text,
                    'row_index': dankoffer['row_index'],
                    'total_count': dankoffer['total_count'],
                    'already_assigned': dankoffer.get('already_assigned', False)
                }
            else:
                preview['alerts']['not_found'].append('Geen Dankoffer verzen gevonden')
        except Exception as e:
            preview['alerts']['warnings'].append(f'Fout bij ophalen Dankoffer: {str(e)}')
        
        return jsonify(preview)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Liturgie Auto-Fill Working File Endpoint (Dropbox Direct)
# ---------------------------------------------------------------------------

WORKING_FILE_PATH = '/working folder/file mingguan/Main Liturgy file.xlsx'

@app.route('/liturgie/auto-fill-working-file', methods=['POST'])
def auto_fill_working_file():
    """Auto-fill the Main Liturgy file from Dropbox working folder"""
    from io import BytesIO
    from openpyxl import load_workbook
    
    try:
        # Get selected changes from request
        data = request.get_json() or {}
        selected_changes = data.get('selected_changes', [])
        print(f"[AutoFill] Selected changes: {selected_changes}")
        
        # Get Dropbox client
        dbx = _get_dbx_liturgie()
        if not dbx:
            return jsonify({'error': 'Dropbox niet geconfigureerd'}), 500
        
        # Read Main Liturgy file from Dropbox
        try:
            resp = dbx.files_download(WORKING_FILE_PATH)
            file_bytes = resp[1].content
        except Exception as e:
            return jsonify({'error': f'Kon Main Liturgy file niet laden van Dropbox: {str(e)}'}), 500
        
        # Load Excel
        wb = load_workbook(BytesIO(file_bytes))
        ws_data = wb['Data']
        ws_active = wb.active
        
        # Get service date from Data sheet
        service_date_raw = ws_data['B3'].value
        print(f"[Working File AutoFill] Data!B3 value: {service_date_raw}")
        
        # Parse date
        if isinstance(service_date_raw, str):
            try:
                service_date = datetime.strptime(service_date_raw, '%d-%m-%Y')
            except ValueError:
                try:
                    service_date = datetime.strptime(service_date_raw, '%Y-%m-%d')
                except ValueError:
                    return jsonify({'error': f'Ongeldige datum formaat in D4: {service_date_raw}'}), 400
        else:
            service_date = service_date_raw
        
        # Track what was filled
        alerts = {
            'already_filled': [],
            'auto_populated': [],
            'not_found': []
        }
        
        # Helper to set cell value
        def set_cell_value(ws, row, col, value, field_name, force=False):
            cell = ws.cell(row=row, column=col)
            current_val = str(cell.value).strip() if cell.value else ''
            
            if current_val and current_val.lower() not in ('nan', 'none', ''):
                if force:
                    # Force update - overwrite existing value
                    cell.value = value
                    alerts['auto_populated'].append(f'{field_name}: {value} (overschreven)')
                    return True
                else:
                    alerts['already_filled'].append(f'{field_name}: {current_val}')
                    return False
            elif value:
                cell.value = value
                alerts['auto_populated'].append(f'{field_name}: {value}')
                return True
            return False
        
        # 1. Populate B4-B12 from Takenrooster
        print(f"[AutoFill] Getting takenrooster for date: {service_date}")
        takenrooster_data = _get_takenrooster()
        takenrooster = takenrooster_data.get('entries', [])
        print(f"[AutoFill] Found {len(takenrooster)} takenrooster entries")
        entry = None
        for tr_entry in takenrooster:
            entry_date = tr_entry.get('date')
            if entry_date and entry_date.date() == service_date.date():
                entry = tr_entry
                break
        
        if entry:
            print(f"[AutoFill] Found entry: {entry}")
            field_mapping = [
                (4, 'Voorganger', 'predikant'),
                (5, 'OvD', 'ovd'),
                (6, '1e Ontvangst', '1eo'),
                (7, '2e Ontvangst', '2eo'),
                (8, 'Voorzangers', 'voorzangers'),
                (9, 'Beamer', 'beamer'),
                (10, 'Geluid', 'multimedia'),
                (11, 'KND', 'knd'),
                (12, 'Tieners', 'tieners'),
            ]
            
            for row, field_name, taken_key in field_mapping:
                new_val = str(entry.get(taken_key, '')).strip()
                print(f"[AutoFill] {field_name} ({taken_key}): '{new_val}'")
                # Force update if field is in selected_changes
                force_update = field_name in selected_changes
                if new_val:
                    result = set_cell_value(ws_active, row, 2, new_val, field_name, force=force_update)
                    print(f"[AutoFill]   -> set_cell_value result: {result} (force={force_update})")
                else:
                    alerts['not_found'].append(f'{field_name}: niet gevonden in takenrooster')
        else:
            print(f"[AutoFill] No entry found for date {service_date}")
            alerts['not_found'].append(f'Geen dienst gevonden in takenrooster voor {service_date.strftime("%d-%m-%Y")}')
        
        # 2. Populate Tikkie link
        try:
            reader = OutlookCollecteReader()
            if reader.is_authenticated():
                email_data = reader.fetch_collecte_data(target_date=service_date, since_days=60)
                tikkie_url = email_data.get('dankoffer_url', '')
                if tikkie_url:
                    tikkie_row = None
                    for row in range(1, ws_active.max_row + 1):
                        label = str(ws_active.cell(row=row, column=1).value).strip().lower() if ws_active.cell(row=row, column=1).value else ''
                        if 'tikkie' in label or 'qr_link' in label:
                            tikkie_row = row
                            break
                    if tikkie_row:
                        set_cell_value(ws_active, tikkie_row, 2, tikkie_url, 'Tikkie link')
                else:
                    alerts['not_found'].append('Tikkie link niet gevonden in e-mail')
            else:
                alerts['not_found'].append('Outlook niet geauthenticeerd - Tikkie link niet opgehaald')
        except Exception as e:
            alerts['not_found'].append(f'Tikkie link ophalen mislukt: {str(e)}')
        
        # 3. Populate Dankoffer verse
        dankoffer = _get_dankoffer_verse(dbx, service_date, mark_as_used=True)
        dankoffer_info = None
        
        if dankoffer:
            dankoffer_row = 21
            dankoffer_book = dankoffer['book']
            
            def set_dankoffer_cell(ws, row, col, value):
                cell = ws.cell(row=row, column=col)
                current_val = str(cell.value).strip() if cell.value else ''
                if current_val and current_val.lower() not in ('nan', 'none', ''):
                    return False
                elif value:
                    cell.value = value
                    return True
                return False
            
            set_dankoffer_cell(ws_active, dankoffer_row, 2, dankoffer_book)
            set_dankoffer_cell(ws_active, dankoffer_row, 3, dankoffer['chapter'])
            verse_text = dankoffer['verse_start']
            if dankoffer['verse_end']:
                verse_text += f'-{dankoffer["verse_end"]}'
            set_dankoffer_cell(ws_active, dankoffer_row, 4, verse_text)
            
            alerts['auto_populated'].append(f'Dankoffer vers: {dankoffer["full_text"]}')
            
            # Build status detail
            status_detail = f'Rij {dankoffer["row_index"]} van {dankoffer["total_count"]} ({dankoffer["unused_count"]} resterend)'
            if dankoffer.get('already_assigned'):
                status_detail = f'Dit vers was al toegewezen aan deze datum • {status_detail}'
            elif dankoffer.get('reset_needed'):
                status_detail = f'Alle verzen waren gebruikt → oudste verzen worden hergebruikt • {status_detail}'
            else:
                status_detail = f'Nieuwe toewijzing toegevoegd aan Dankoffer.xlsx • {status_detail}'
            
            dankoffer_info = {
                'verse': dankoffer['full_text'],
                'row_index': dankoffer['row_index'],
                'total_count': dankoffer['total_count'],
                'unused_count': dankoffer['unused_count'],
                'already_assigned': dankoffer.get('already_assigned', False),
                'reset_needed': dankoffer.get('reset_needed', False),
                'marked_as_used': dankoffer.get('marked_as_used', True),
                'status_detail': status_detail
            }
        else:
            alerts['not_found'].append('Geen Dankoffer verzen gevonden')
        
        # Save back to Dropbox
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        try:
            dbx.files_upload(
                output.getvalue(),
                WORKING_FILE_PATH,
                mode=dropbox.files.WriteMode.overwrite
            )
        except Exception as e:
            return jsonify({'error': f'Kon bestand niet opslaan naar Dropbox: {str(e)}'}), 500
        
        return jsonify({
            'success': True,
            'service_date': service_date.strftime('%d-%m-%Y'),
            'alerts': alerts,
            'dankoffer_info': dankoffer_info,
            'message': f'Bestand bijgewerkt op Dropbox: {WORKING_FILE_PATH}'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/liturgie/preview-working-file', methods=['GET'])
def preview_working_file():
    """
    Preview what changes would be made to Main Liturgy file.xlsx in Dropbox working folder.
    """
    from openpyxl import load_workbook
    from io import BytesIO
    
    try:
        dbx = _get_dbx_liturgie()
        if not dbx:
            return jsonify({'error': 'Dropbox niet geconfigureerd'}), 500
        
        # Read file
        try:
            resp = dbx.files_download(WORKING_FILE_PATH)
            excel_bytes = resp[1].content
        except Exception as e:
            return jsonify({'error': f'Kon bestand niet laden: {str(e)}'}), 500
        
        wb = load_workbook(BytesIO(excel_bytes))
        
        # Get service date from Data sheet cell B3
        try:
            ws_data = wb['Data']
        except KeyError:
            return jsonify({'error': 'Data tab niet gevonden in Excel bestand'}), 400
            
        service_date_raw = ws_data.cell(row=3, column=2).value  # B3
        print(f'[Working File Preview] Data!B3 value: {repr(service_date_raw)}')
        if not service_date_raw:
            return jsonify({'error': f'Geen datum gevonden in Data!B3 (waarde: {repr(service_date_raw)})'}), 400
        
        # Use active sheet for reading current values
        ws = wb.active
            
        if isinstance(service_date_raw, str):
            try:
                service_date = datetime.strptime(service_date_raw, '%d-%m-%Y')
            except ValueError:
                service_date = datetime.strptime(service_date_raw, '%Y-%m-%d')
        else:
            service_date = service_date_raw
        
        # Build preview
        preview = {
            'service_date': service_date.strftime('%d-%m-%Y'),
            'current_values': {},
            'proposed_changes': {},
            'alerts': {'not_found': [], 'warnings': []}
        }
        
        # Current values
        field_rows = {
            'Voorganger': 4, 'OvD': 5, '1e Ontvangst': 6, '2e Ontvangst': 7,
            'Voorzangers': 8, 'Beamer': 9, 'Geluid': 10, 'KND': 11, 'Tieners': 12
        }
        
        for field, row in field_rows.items():
            val = str(ws.cell(row=row, column=2).value).strip() if ws.cell(row=row, column=2).value else ''
            preview['current_values'][field] = val
        
        # Tikkie current
        for row in range(1, ws.max_row + 1):
            label = str(ws.cell(row=row, column=1).value).strip().lower() if ws.cell(row=row, column=1).value else ''
            if 'tikkie' in label or 'qr_link' in label:
                val = str(ws.cell(row=row, column=2).value).strip() if ws.cell(row=row, column=2).value else ''
                preview['current_values']['Tikkie link'] = val
                break
        
        # Dankoffer current
        b21 = str(ws.cell(row=21, column=2).value).strip() if ws.cell(row=21, column=2).value else ''
        c21 = str(ws.cell(row=21, column=3).value).strip() if ws.cell(row=21, column=3).value else ''
        d21 = str(ws.cell(row=21, column=4).value).strip() if ws.cell(row=21, column=4).value else ''
        if b21 and c21 and d21:
            preview['current_values']['Dankoffer vers'] = f'{b21} {c21}:{d21}'
        else:
            preview['current_values']['Dankoffer vers'] = ''
        
        # Get proposed values from sources
        try:
            takenrooster_data = _get_takenrooster()
            takenrooster = takenrooster_data.get('entries', [])
            entry = None
            for tr_entry in takenrooster:
                entry_date = tr_entry.get('date')
                if entry_date and entry_date.date() == service_date.date():
                    entry = tr_entry
                    break
            if entry:
                print(f"[DEBUG] Takenrooster entry keys: {list(entry.keys())}")
                for field, key in [('Voorganger', 'predikant'), ('OvD', 'ovd'), ('1e Ontvangst', '1eo'),
                                   ('2e Ontvangst', '2eo'), ('Voorzangers', 'voorzangers'), 
                                   ('Beamer', 'beamer'), ('Geluid', 'multimedia'),
                                   ('KND', 'knd'), ('Tieners', 'tieners')]:
                    val = str(entry.get(key, '')).strip()
                    print(f"[DEBUG] {field} ({key}): '{val}'")
                    if val:
                        preview['proposed_changes'][field] = val
            else:
                preview['alerts']['not_found'].append('Geen dienst in takenrooster')
        except Exception as e:
            preview['alerts']['warnings'].append(f'Takenrooster fout: {str(e)}')
        
        try:
            reader = OutlookCollecteReader()
            if reader.is_authenticated():
                email_data = reader.fetch_collecte_data(target_date=service_date, since_days=60)
                url = email_data.get('dankoffer_url', '')
                if url:
                    preview['proposed_changes']['Tikkie link'] = url
                else:
                    preview['alerts']['not_found'].append('Tikkie niet gevonden')
            else:
                preview['alerts']['warnings'].append('Outlook niet geauthenticeerd')
        except Exception as e:
            preview['alerts']['warnings'].append(f'Tikkie fout: {str(e)}')
        
        try:
            dankoffer = _get_dankoffer_verse(dbx, service_date, mark_as_used=False)
            if dankoffer:
                verse = f"{dankoffer['book']} {dankoffer['chapter']}:{dankoffer['verse_start']}"
                if dankoffer['verse_end']:
                    verse += f"-{dankoffer['verse_end']}"
                preview['proposed_changes']['Dankoffer vers'] = verse
            else:
                preview['alerts']['not_found'].append('Dankoffer niet gevonden')
        except Exception as e:
            preview['alerts']['warnings'].append(f'Dankoffer fout: {str(e)}')
        
        return jsonify(preview)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500




# =============================================================================
# Sender Campaign Routes - Updated 2026-05-14
# =============================================================================

@app.route('/fetch-ole-data', methods=['POST'])
@_password_required
def fetch_ole_data():
    """Fetch OLE preekroster data for a given date. Deployed: 2026-05-14 20:25"""
    data = request.get_json() or {}
    iso_date = data.get('date', '')
    print(f"[OLE Fetch] Received request for date: {iso_date}")

    if not iso_date:
        print("[OLE Fetch] Error: no date provided")
        return jsonify({'error': 'no date provided'}), 400

    try:
        selected_date = datetime.strptime(iso_date, '%Y-%m-%d')
        print(f"[OLE Fetch] Parsed date: {selected_date}")
    except ValueError as e:
        print(f"[OLE Fetch] Error parsing date: {e}")
        return jsonify({'error': 'invalid date format'}), 400

    try:
        from data_sources.preekroster_scraper import PreekrosterScraper
        scraper = PreekrosterScraper()
        print("[OLE Fetch] Scraper initialized - fixed import")

        # Fetch OLE data for the selected date
        ole_data = scraper.get_ole_service_for_date(selected_date)
        print(f"[OLE Fetch] OLE data retrieved: {ole_data}")

        # Also try to get QR code and Tikkie URL from recent emails
        qr_filename = None
        ole_url = None

        try:
            from data_sources.email_reader import EmailReader
            reader = EmailReader()
            print("[OLE Fetch] Email reader initialized")
            # Fetch collecte data for the selected date
            collecte_data = reader.fetch_collecte_data(target_date=selected_date)
            print(f"[OLE Fetch] Collecte data: {collecte_data}")
            if collecte_data:
                qr_filename = collecte_data.get('ole_qr') or collecte_data.get('dankoffer_qr')
                ole_url = collecte_data.get('ole_url') or collecte_data.get('dankoffer_url')
                print(f"[OLE Fetch] QR filename: {qr_filename}, URL: {ole_url}")
        except Exception as e:
            print(f"[OLE Fetch] Email fetch error: {e}")
            import traceback
            traceback.print_exc()

        # Translate location code to full name
        location_code = ole_data.get('location', '')
        location_names = {
            'AM': 'Kerkgebouw in Amstelveen',
            'DH': 'Kerkgebouw in Den Haag',
            'TB': 'Pauluskerk te Tilburg'
        }
        location_full = location_names.get(location_code, location_code)

        # Fetch thema, bijbeltekst, youtube, liturgie and collecte from GKIN website
        thema = ''
        bible_verse = ''
        youtube_link = ''
        liturgie_url = ''
        qr_image_b64 = ''
        collecte_ovv = ''
        try:
            from data_sources.gkin_ole_scraper import GKINOLEScraper
            scraper = GKINOLEScraper()
            web_data = scraper.fetch_for_date(selected_date)
            if not web_data.get('not_found'):
                thema = web_data.get('thema', '')
                bible_verse = web_data.get('bible_verse', '')
                youtube_link = web_data.get('youtube_link', '')
                liturgie_url = web_data.get('liturgie_url', '')
                qr_image_b64 = web_data.get('qr_image_b64', '')
                # QR from website overrides local file
                if web_data.get('qr_image_url'):
                    qr_filename = web_data['qr_image_url']
                # Override predikant/location/time from website if better
                if web_data.get('predikant') and not ole_data.get('predikant'):
                    ole_data['predikant'] = web_data['predikant']
                if web_data.get('location_code'):
                    location_code = web_data['location_code']
                    location_full = web_data.get('location', location_full)
                if web_data.get('time'):
                    ole_data['time'] = web_data['time']
                if web_data.get('collecte_url') and not ole_url:
                    ole_url = web_data['collecte_url']
                collecte_ovv = web_data.get('collecte_ovv', '')
                print(f"[OLE Fetch] Website data: thema={thema!r}, bible={bible_verse!r}, yt={youtube_link!r}, liturgie={liturgie_url!r}, qr_b64={bool(qr_image_b64)}, ovv={collecte_ovv!r}")
            else:
                print(f"[OLE Fetch] No OLE article found on GKIN website for {selected_date.strftime('%d-%m-%Y')}")
        except Exception as e:
            print(f"[OLE Fetch] Website scrape error: {e}")
            import traceback; traceback.print_exc()

        result = {
            'ole_predikant': ole_data.get('predikant', ''),
            'ole_location_code': location_code,
            'ole_location': location_full,
            'ole_time': ole_data.get('time', '10:00'),
            'ole_qr': qr_filename,
            'ole_qr_b64': qr_image_b64,
            'ole_url': ole_url,
            'ole_thema': thema,
            'ole_bible_verse': bible_verse,
            'ole_youtube_link': youtube_link,
            'ole_liturgie_url': liturgie_url,
            'ole_collecte_ovv': collecte_ovv,
            'success': True
        }
        print(f"[OLE Fetch] Returning result: {result}")
        return jsonify(result)

    except Exception as e:
        print(f"[OLE Fetch] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/campaign')
@_password_required
def campaign_index():
    """Render the campaign template selection page."""
    return render_template('campaign_select.html')


@app.route('/campaign/ole')
@_password_required
def campaign_ole():
    """Render the OLE Sender campaign generator page."""
    taken = _get_takenrooster()
    dutch_months = ['januari', 'februari', 'maart', 'april', 'mei', 'juni',
                    'juli', 'augustus', 'september', 'oktober', 'november', 'december']
    dutch_days = ['maandag', 'dinsdag', 'woensdag', 'donderdag', 'vrijdag', 'zaterdag', 'zondag']
    today = datetime.now().date()
    dates = []
    for entry in taken['entries']:
        d = entry['date']
        d_date = d.date() if hasattr(d, 'date') else d
        if d_date < today:
            continue
        label = f"{dutch_days[d.weekday()]} {d.day} {dutch_months[d.month - 1]} {d.year}"
        dates.append({
            'value': d.strftime('%Y-%m-%d'),
            'label': label,
            'day_idx':   d.weekday(),
            'month_idx': d.month - 1,
            'day_num':   d.day,
            'year':      d.year,
            'suffix':    '',
            'predikant': '',
            'ovd': entry.get('ovd', ''),
            'beamer': entry.get('beamer', ''),
            'voorzangers': entry.get('voorzangers', ''),
        })
    return render_template('campaign.html', dates=dates)


@app.route('/campaign/preview', methods=['POST'])
@_password_required
def campaign_preview():
    """Generate campaign preview."""
    from sender_campaign import SenderCampaignGenerator
    
    data = request.get_json() or {}
    iso_date = data.get('date', '')
    if not iso_date:
        return jsonify({'error': 'no date'}), 400
    
    try:
        selected_date = datetime.strptime(iso_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date format'}), 400
    
    try:
        taken = _get_takenrooster()
        entry = None
        for e in taken['entries']:
            e_date = e['date'].date() if hasattr(e['date'], 'date') else e['date']
            if e_date == selected_date.date():
                entry = e
                break
        
        if not entry:
            return jsonify({'error': f'No rooster entry for {iso_date}'}), 404
        
        reader = DropboxExcelReader()
        meded = reader.get_mededelingen(mededelingen_date=selected_date)
        
        # OLE fields
        theme = data.get('theme', '')
        bible_verse = data.get('bible_verse', '')
        youtube_link = data.get('youtube_link', '')
        liturgie_url = data.get('liturgie_url', '')
        collecte_url = data.get('collecte_url', '')
        qr_image_url = data.get('qr_image_url', '')
        ole_location = data.get('ole_location', '')
        ole_time = data.get('ole_time', '10:00')
        ole_predikant = data.get('ole_predikant', '')
        collecte_ovv = data.get('collecte_ovv', '')
        
        predikant_to_use = ole_predikant if ole_predikant else entry.get('predikant', '')
        
        generator = SenderCampaignGenerator()
        html_content = generator.generate_html(
            service_date=selected_date,
            predikant=predikant_to_use,
            theme=theme,
            bible_verse=bible_verse,
            youtube_link=youtube_link,
            liturgie_url=liturgie_url,
            collecte_url=collecte_url,
            qr_image_url=qr_image_url,
            ole_location=ole_location,
            ole_time=ole_time,
            collecte_ovv=collecte_ovv
        )
        
        nl_months = ['januari', 'februari', 'maart', 'april', 'mei', 'juni',
                     'juli', 'augustus', 'september', 'oktober', 'november', 'december']
        nl_days = ['maandag', 'dinsdag', 'woensdag', 'donderdag', 'vrijdag', 'zaterdag', 'zondag']
        date_str = f"{selected_date.day} {nl_months[selected_date.month - 1]} {selected_date.year}"
        day_name = nl_days[selected_date.weekday()]
        time_clean = (ole_time if ole_time else '10:00').replace('u', '').replace('U', '')
        
        return jsonify({
            'date': date_str,
            'predikant': predikant_to_use,
            'location': ole_location,
            'time': time_clean,
            'subject': f"GKIN (OLE): Online Landelijke Eredienst {day_name} {date_str}, {time_clean}u",
            'html_preview': html_content,
            'success': True
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/campaign/lists', methods=['GET'])
@_password_required
def campaign_lists():
    """Return available Sender subscriber lists."""
    from sender_campaign import SenderCampaignGenerator
    try:
        gen = SenderCampaignGenerator()
        lists = gen.get_lists()
        print(f"[Campaign Lists] Raw groups: {lists[:2] if lists else lists}")
        return jsonify({'success': True, 'lists': lists})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/campaign/create', methods=['POST'])
@_password_required
def campaign_create():
    """Create the Sender campaign."""
    from sender_campaign import SenderCampaignGenerator
    
    data = request.get_json() or {}
    iso_date = data.get('date', '')
    subject = data.get('subject', '')
    name = data.get('name', '')
    scheduled_at_input = data.get('scheduled_at', None)
    list_ids = data.get('list_ids', None)
    
    if not iso_date:
        return jsonify({'error': 'no date'}), 400
    
    try:
        selected_date = datetime.strptime(iso_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date format'}), 400
    
    try:
        taken = _get_takenrooster()
        entry = None
        for e in taken['entries']:
            e_date = e['date'].date() if hasattr(e['date'], 'date') else e['date']
            if e_date == selected_date.date():
                entry = e
                break
        
        if not entry:
            return jsonify({'error': f'No rooster entry for {iso_date}'}), 404
        
        reader = DropboxExcelReader()
        meded = reader.get_mededelingen(mededelingen_date=selected_date)
        
        # OLE fields
        theme = data.get('theme', '')
        bible_verse = data.get('bible_verse', '')
        youtube_link = data.get('youtube_link', '')
        liturgie_url = data.get('liturgie_url', '')
        collecte_url = data.get('collecte_url', '')
        qr_image_url = data.get('qr_image_url', '')
        ole_location = data.get('ole_location', '')
        ole_time = data.get('ole_time', '10:00')
        ole_predikant = data.get('ole_predikant', '')
        collecte_ovv = data.get('collecte_ovv', '')
        
        predikant_to_use = ole_predikant if ole_predikant else entry.get('predikant', '')
        
        generator = SenderCampaignGenerator()
        html_content = generator.generate_html(
            service_date=selected_date,
            predikant=predikant_to_use,
            theme=theme,
            bible_verse=bible_verse,
            youtube_link=youtube_link,
            liturgie_url=liturgie_url,
            collecte_url=collecte_url,
            qr_image_url=qr_image_url,
            ole_location=ole_location,
            ole_time=ole_time,
            collecte_ovv=collecte_ovv
        )
        
        nl_months = ['januari', 'februari', 'maart', 'april', 'mei', 'juni',
                     'juli', 'augustus', 'september', 'oktober', 'november', 'december']
        nl_days = ['maandag', 'dinsdag', 'woensdag', 'donderdag', 'vrijdag', 'zaterdag', 'zondag']
        date_str = f"{selected_date.day} {nl_months[selected_date.month - 1]} {selected_date.year}"
        day_name = nl_days[selected_date.weekday()]
        time_clean = (ole_time if ole_time else '10:00').replace('u', '').replace('U', '')
        
        # Use scheduled_at from frontend if provided
        scheduled_at = scheduled_at_input if scheduled_at_input else None
        
        result = generator.create_campaign(
            name=name or f"OLE {selected_date.strftime('%d-%m-%Y')}",
            subject=subject or f"GKIN (OLE): Online Landelijke Eredienst {day_name} {date_str}, {time_clean}u",
            html_content=html_content,
            scheduled_at=scheduled_at,
            list_ids=list_ids if list_ids else None
        )
        
        if 'error' in result:
            error_msg = result['error']
            if result.get('details'):
                import json
                details = result['details']
                error_msg += f" - Details: {json.dumps(details)}"
            return jsonify({'success': False, 'error': error_msg}), 500
        
        return jsonify({
            'success': True,
            'campaign_id': result.get('data', {}).get('id', 'unknown'),
            'name': name,
            'subject': subject,
            'scheduled': result.get('scheduled', False),
            'schedule_time': result.get('schedule_time', ''),
            'schedule_error': result.get('schedule_error', '')
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/upload-to-sender', methods=['POST'])
@_password_required
def upload_to_sender():
    """Upload file to Sender."""
    from sender_campaign import SenderCampaignGenerator
    
    data = request.get_json() or {}
    local_path = data.get('local_path', '')
    
    if not local_path:
        return jsonify({'success': False, 'error': 'No path provided'}), 400
    
    if local_path.startswith('/uploads/'):
        filename = local_path.replace('/uploads/', '')
        full_path = os.path.join(UPLOAD_DIR, filename)
    elif os.path.exists(local_path):
        full_path = local_path
    else:
        full_path = os.path.join(UPLOAD_DIR, local_path)
    
    if not os.path.exists(full_path):
        return jsonify({'success': False, 'error': f'File not found: {full_path}'}), 404
    
    try:
        generator = SenderCampaignGenerator()
        result = generator.upload_file(full_path)
        
        if result.get('success'):
            return jsonify({'success': True, 'url': result['url'], 'file_id': result['file_id']})
        else:
            return jsonify({'success': False, 'error': result.get('error', 'Upload failed')}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/campaign/upload-liturgie', methods=['POST'])
@_password_required
def campaign_upload_liturgie():
    """Upload a liturgie file to Dropbox and return a public shared link URL."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'Empty filename'}), 400
    import werkzeug.utils
    safe_name = werkzeug.utils.secure_filename(f.filename)
    file_bytes = f.read()
    try:
        import dropbox as _dropbox
        from dropbox.exceptions import ApiError as _DbxApiError
        dbx = _dropbox.Dropbox(
            oauth2_refresh_token=DROPBOX_REFRESH_L,
            app_key=DROPBOX_APP_KEY_L,
            app_secret=DROPBOX_APP_SECRET_L,
        )
        dropbox_path = f"/#Kerkbode GKIN Amstelveen/OLE-Liturgie/{safe_name}"
        dbx.files_upload(file_bytes, dropbox_path, mode=_dropbox.files.WriteMode.overwrite)
        # Get or create shared link (legacy API, no sharing.write scope needed)
        try:
            link_meta = dbx.sharing_create_shared_link(dropbox_path)
        except _DbxApiError:
            links = dbx.sharing_list_shared_links(path=dropbox_path, direct_only=True)
            link_meta = links.links[0] if links.links else None
        if not link_meta:
            return jsonify({'success': False, 'error': 'Could not create Dropbox shared link'}), 500
        raw = link_meta.url
        url = raw.replace('www.dropbox.com', 'dl.dropboxusercontent.com').replace('?dl=0', '').replace('?dl=1', '') + '?dl=1'
        return jsonify({'success': True, 'url': url})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _normalize_salutation(name: str) -> str:
    """Lowercase the salutation prefix (ds./zr./br./mw./dhr./dr.) in a predikant name."""
    import re as _re
    return _re.sub(
        r'^(Ds|Zr|Br|Mw|Dhr|Dr)\.',
        lambda m: m.group(0).lower(),
        (name or '').strip(),
        flags=_re.IGNORECASE
    )


@app.route('/fetch-pm-data', methods=['POST'])
@_password_required
def fetch_pm_data():
    """Fetch AM predikant, OLE location/predikant, youtube link and preek URL for a given date."""
    data = request.get_json() or {}
    iso_date = data.get('date', '')
    if not iso_date:
        return jsonify({'error': 'no date provided'}), 400
    try:
        selected_date = datetime.strptime(iso_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date format'}), 400
    try:
        from data_sources.preekroster_scraper import PreekrosterScraper
        from data_sources.gkin_ole_scraper import GKINOLEScraper
        scraper = PreekrosterScraper()
        # AM predikant — from takenrooster
        taken = _get_takenrooster()
        am_predikant = ''
        for entry in taken['entries']:
            e_date = entry['date'].date() if hasattr(entry['date'], 'date') else entry['date']
            if e_date == selected_date.date():
                am_predikant = _normalize_salutation(entry.get('predikant', ''))
                break
        # OLE location + predikant — from preekroster
        ole_data = scraper.get_ole_service_for_date(selected_date)
        location_code = ole_data.get('location', '').strip().upper()
        ole_predikant = _normalize_salutation(ole_data.get('predikant', ''))
        # YouTube link + preek URL — from gkin.org website
        youtube_link = ''
        preek_url = ''
        try:
            web = GKINOLEScraper().fetch_for_date(selected_date)
            if web and not web.get('not_found'):
                youtube_link = web.get('youtube_link', '')
                preek_url = web.get('preek_url', '')
                # Also use website predikant if preekroster didn't return one
                if not ole_predikant and web.get('predikant'):
                    ole_predikant = _normalize_salutation(web.get('predikant', ''))
        except Exception as web_err:
            print(f'[fetch-pm-data] Website fetch error: {web_err}')
        return jsonify({
            'success': True,
            'am_predikant': am_predikant,
            'ole_location': location_code,
            'ole_predikant': ole_predikant,
            'ole_time': ole_data.get('time', ''),
            'youtube_link': youtube_link,
            'preek_ole_url': preek_url,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/campaign/pm')
@_password_required
def campaign_pm():
    """Render the Preek & Mededelingen Sender campaign generator page."""
    taken = _get_takenrooster()
    dutch_months = ['januari', 'februari', 'maart', 'april', 'mei', 'juni',
                    'juli', 'augustus', 'september', 'oktober', 'november', 'december']
    dutch_days = ['maandag', 'dinsdag', 'woensdag', 'donderdag', 'vrijdag', 'zaterdag', 'zondag']
    today = datetime.now().date()
    dates = []
    for entry in taken['entries']:
        d = entry['date']
        d_date = d.date() if hasattr(d, 'date') else d
        if d_date < today:
            continue
        label = f"{dutch_days[d.weekday()]} {d.day} {dutch_months[d.month - 1]} {d.year}"
        dates.append({
            'value': d.strftime('%Y-%m-%d'),
            'label': label,
            'day_idx':   d.weekday(),
            'month_idx': d.month - 1,
            'day_num':   d.day,
            'year':      d.year,
            'suffix':    '',
            'predikant': entry.get('predikant', ''),
            'ovd': entry.get('ovd', ''),
            'beamer': entry.get('beamer', ''),
            'voorzangers': entry.get('voorzangers', ''),
        })
    return render_template('campaign_pm.html', dates=dates)


@app.route('/campaign/pm/preview', methods=['POST'])
@_password_required
def campaign_pm_preview():
    """Generate Preek & Mededelingen campaign preview."""
    from sender_campaign import SenderCampaignGenerator
    data = request.get_json() or {}
    iso_date = data.get('date', '')
    if not iso_date:
        return jsonify({'error': 'no date'}), 400
    try:
        selected_date = datetime.strptime(iso_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date format'}), 400
    try:
        generator = SenderCampaignGenerator()
        html_content = generator.generate_pm_html(
            service_date=selected_date,
            am_predikant=data.get('am_predikant', ''),
            mededelingen_url=data.get('mededelingen_url', ''),
            preek_am_url=data.get('preek_am_url', ''),
            ole_location=data.get('ole_location', ''),
            ole_predikant=data.get('ole_predikant', ''),
            youtube_link=data.get('youtube_link', ''),
            preek_ole_url=data.get('preek_ole_url', ''),
        )
        return jsonify({'success': True, 'html_preview': html_content})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/campaign/pm/create', methods=['POST'])
@_password_required
def campaign_pm_create():
    """Create the Preek & Mededelingen Sender campaign."""
    from sender_campaign import SenderCampaignGenerator
    data = request.get_json() or {}
    iso_date = data.get('date', '')
    subject = data.get('subject', '')
    name = data.get('name', '')
    scheduled_at = data.get('scheduled_at', None)
    list_ids = data.get('list_ids', None)
    if not iso_date:
        return jsonify({'error': 'no date'}), 400
    try:
        selected_date = datetime.strptime(iso_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'invalid date format'}), 400
    try:
        nl_months = ['januari', 'februari', 'maart', 'april', 'mei', 'juni',
                     'juli', 'augustus', 'september', 'oktober', 'november', 'december']
        date_str = f"{selected_date.day} {nl_months[selected_date.month - 1]} {selected_date.year}"
        generator = SenderCampaignGenerator()
        html_content = generator.generate_pm_html(
            service_date=selected_date,
            am_predikant=data.get('am_predikant', ''),
            mededelingen_url=data.get('mededelingen_url', ''),
            preek_am_url=data.get('preek_am_url', ''),
            ole_location=data.get('ole_location', ''),
            ole_predikant=data.get('ole_predikant', ''),
            youtube_link=data.get('youtube_link', ''),
            preek_ole_url=data.get('preek_ole_url', ''),
        )
        result = generator.create_campaign(
            name=name or f"PM {selected_date.strftime('%d-%m-%Y')}",
            subject=subject or f"GKIN Amstelveen {date_str}: Preek & mededelingen",
            html_content=html_content,
            scheduled_at=scheduled_at,
            list_ids=list_ids if list_ids else None,
        )
        if 'error' in result:
            error_msg = result['error']
            if result.get('details'):
                import json as _json
                error_msg += f" - Details: {_json.dumps(result['details'])}"
            return jsonify({'success': False, 'error': error_msg}), 500
        return jsonify({
            'success': True,
            'campaign_id': result.get('data', {}).get('id', 'unknown'),
            'name': name,
            'subject': subject,
            'scheduled': result.get('scheduled', False),
            'schedule_time': result.get('schedule_time', ''),
            'schedule_error': result.get('schedule_error', ''),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# =============================================================================
# Health Check
# =============================================================================

@app.route('/health')
def health_check():
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()})


# =============================================================================
# Main Entry Point
# =============================================================================

if __name__ == '__main__':
    app.run(debug=True, port=5000)
