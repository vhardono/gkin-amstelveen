"""
Dropbox Excel Reader
Reads Takenrooster Excel file from Dropbox for church bulletin data.
Extracts dates, predikant, and OvD (with full name lookup from People tab).
"""

import os
import json
import re
import pandas as pd
import dropbox
from dropbox.exceptions import ApiError
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Any, Optional
from config import Config

TAKENROOSTER_PATH = '/# Kerkbode GKIN Amstelveen/Rooster/Takenrooster_GKIN_Amstelveen_2026.xlsx'
MEDEDELINGEN_PATH_TEMPLATE = '/# Kerkbode GKIN Amstelveen/{year}/Mededelingen Overzicht.xlsx'

MEDEDELINGEN_IMG_DIR = os.getenv('MEDEDELINGEN_IMAGES_DIR',
                                 '/data/mededelingen_images' if os.path.isdir('/data') else './mededelingen_images')
MEDEDELINGEN_IMG_META = os.path.join(MEDEDELINGEN_IMG_DIR, 'images.json')


def _mededelingen_image_dir_for_date(d: datetime) -> str:
    return os.path.join(MEDEDELINGEN_IMG_DIR, str(d.year), d.strftime('%Y%m%d'))


def _mededelingen_image_meta() -> dict:
    if not os.path.exists(MEDEDELINGEN_IMG_META):
        return {}
    try:
        with open(MEDEDELINGEN_IMG_META, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _save_mededelingen_image_meta(meta: dict) -> None:
    os.makedirs(MEDEDELINGEN_IMG_DIR, exist_ok=True)
    with open(MEDEDELINGEN_IMG_META, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


def _mededelingen_image_url_for_date(d: datetime) -> Optional[str]:
    meta = _mededelingen_image_meta()
    key = f"{d.year}-{d.strftime('%Y%m%d')}"
    rec = meta.get(key)
    if not rec:
        return None
    return f"/mededelingen-image/{d.year}/{d.strftime('%Y%m%d')}/{rec['filename']}"


class DropboxExcelReader:
    def __init__(self):
        """Initialize Dropbox client using refresh token for long-lived access"""
        if not Config.DROPBOX_REFRESH_TOKEN:
            raise ValueError("DROPBOX_REFRESH_TOKEN not configured")

        self.dbx = dropbox.Dropbox(
            oauth2_refresh_token=Config.DROPBOX_REFRESH_TOKEN,
            app_key=Config.DROPBOX_APP_KEY,
            app_secret=Config.DROPBOX_APP_SECRET,
        )
        self._people_map: Optional[Dict[str, Dict]] = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def get_takenrooster(self) -> Dict[str, Any]:
        """Read the Takenrooster and return structured data.

        Returns dict with:
            'dates'   – list of available service dates (datetime objects)
            'entries' – list of dicts with keys: date, dag, predikant, ovd, opmerking
                        where ovd is the full name with salutation from People tab.
        """
        try:
            _, response = self.dbx.files_download(TAKENROOSTER_PATH)
            content = BytesIO(response.content)

            people_df = pd.read_excel(content, sheet_name='People', header=0)
            self._build_people_map(people_df)

            content.seek(0)
            current_df = pd.read_excel(content, sheet_name='CURRENT', header=None)

            entries = self._parse_current_sheet(current_df)
            dates = [e['date'] for e in entries]

            print(f"Takenrooster: {len(entries)} services loaded")
            return {
                'source': TAKENROOSTER_PATH,
                'dates': dates,
                'entries': entries,
            }

        except Exception as e:
            print(f"Error reading takenrooster: {e}")
            return {'error': str(e), 'dates': [], 'entries': []}

    def get_mededelingen(self, mededelingen_date: datetime = None) -> Dict[str, Any]:
        """Read Mededelingen Overzicht from the Output tab.

        The year in the file path is derived from the selected mededelingen_date.

        Returns dict with:
            'regionale_nl'  – Regionale Mededelingen (Nederlands) from B2
            'landelijke_nl' – Landelijke Mededelingen (Nederlands) from B3
            'regionale_id'  – Regionale Mededelingen (Bahasa Indonesia) from C2
            'landelijke_id' – Landelijke Mededelingen (Bahasa Indonesia) from C3
        """
        if mededelingen_date is None:
            year = datetime.now().year
        else:
            year = mededelingen_date.year

        path = MEDEDELINGEN_PATH_TEMPLATE.format(year=year)
        try:
            print(f"Reading Mededelingen Overzicht: {path}, date: {mededelingen_date}")
            _, response = self.dbx.files_download(path)

            # Read from Output sheet (contains mededelingen text)
            df_output = pd.read_excel(BytesIO(response.content), sheet_name='Output', header=None)

            def _clean(val):
                if pd.isna(val):
                    return ''
                s = str(val).strip()
                # Treat Excel formula errors as missing so we fall back to computing from the year sheet
                if s.startswith('#') or s.lower() in ('#name?', '#value!', '#ref!', '#num!', '#n/a'):
                    return ''
                return s

            regionale_nl = _clean(df_output.iloc[1, 1])
            landelijke_nl = _clean(df_output.iloc[2, 1])
            regionale_id = _clean(df_output.iloc[1, 2])
            landelijke_id = _clean(df_output.iloc[2, 2])

            # If Output tab is broken/empty, compute active mededelingen from the year sheet
            if not (regionale_nl and landelijke_nl and regionale_id and landelijke_id):
                print('[DropboxReader] Output tab empty or contains errors, computing from year sheet')
                df_year = pd.read_excel(BytesIO(response.content), sheet_name=str(year), header=0)
                # Find the status column (last 'Active' formula column, usually column I)
                status_col = None
                for col in df_year.columns:
                    if df_year[col].astype(str).str.contains('Active|Past|Future', case=False, na=False).any():
                        status_col = col
                if status_col:
                    active = df_year[df_year[status_col] == 'Active']
                    regionale = active[active.iloc[:, 0] == 'Regionale']
                    landelijke = active[active.iloc[:, 0] == 'Landelijke']
                    # Column C = Nederlands (index 2), D = Bahasa Indonesia (index 3)
                    if not regionale_nl:
                        regionale_nl = '\n\n'.join(regionale.iloc[:, 2].dropna().astype(str).tolist())
                    if not landelijke_nl:
                        landelijke_nl = '\n\n'.join(landelijke.iloc[:, 2].dropna().astype(str).tolist())
                    if not regionale_id:
                        regionale_id = '\n\n'.join(regionale.iloc[:, 3].dropna().astype(str).tolist())
                    if not landelijke_id:
                        landelijke_id = '\n\n'.join(landelijke.iloc[:, 3].dropna().astype(str).tolist())

            print(f"Mededelingen loaded: regionale={len(regionale_nl)} chars, landelijke={len(landelijke_nl)} chars")
            return {
                'source': path,
                'regionale_nl': regionale_nl,
                'landelijke_nl': landelijke_nl,
                'regionale_id': regionale_id,
                'landelijke_id': landelijke_id,
            }

        except Exception as e:
            print(f"Error reading mededelingen: {e}")
            return {'error': str(e), 'regionale_nl': '', 'landelijke_nl': '',
                    'regionale_id': '', 'landelijke_id': ''}

    def save_mededelingen_output(self, year: int, regionale_nl: str, regionale_id: str,
                                 landelijke_nl: str, landelijke_id: str) -> Dict[str, Any]:
        """Update the Output sheet of Mededelingen Overzicht and write back to Dropbox."""
        from openpyxl import load_workbook

        path = MEDEDELINGEN_PATH_TEMPLATE.format(year=year)
        try:
            print(f"Writing Mededelingen Overzicht: {path}")
            _, response = self.dbx.files_download(path)

            wb = load_workbook(BytesIO(response.content))
            if 'Output' not in wb.sheetnames:
                return {'success': False, 'error': 'Output sheet not found'}

            ws = wb['Output']
            # Layout: A1='' B1='Details - Nederlands' C1='Details - Bahasa Indonesia'
            # A2='Regionale' B2=nl C2=id
            # A3='Landelijke' B3=nl C3=id
            ws.cell(2, 2).value = regionale_nl
            ws.cell(2, 3).value = regionale_id
            ws.cell(3, 2).value = landelijke_nl
            ws.cell(3, 3).value = landelijke_id

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            self.dbx.files_upload(
                out.read(),
                path,
                mode=dropbox.files.WriteMode.overwrite,
                mute=True
            )

            return {'success': True, 'path': path}

        except Exception as e:
            print(f"Error saving mededelingen: {e}")
            return {'success': False, 'error': str(e)}

    def get_mededelingen_rows(self, year: int) -> List[Dict[str, Any]]:
        """Return all rows from the year sheet as editable dicts.

        Each dict contains:
        row_index, category, type, nl_title, nl_body, id_title, id_body,
        first_date, last_date, event_date, source, status, image_url.
        """
        path = MEDEDELINGEN_PATH_TEMPLATE.format(year=year)
        rows = []
        try:
            _, response = self.dbx.files_download(path)
            df = pd.read_excel(BytesIO(response.content), sheet_name=str(year), header=0)

            for idx, raw in df.iterrows():
                # skip completely empty rows
                if raw.isna().all():
                    continue

                def _cell(i):
                    v = raw.iloc[i] if i < len(raw) else None
                    if pd.isna(v):
                        return ''
                    return str(v).strip()

                def _date_cell(i):
                    v = raw.iloc[i] if i < len(raw) else None
                    if pd.isna(v):
                        return ''
                    if hasattr(v, 'date'):
                        return v.strftime('%Y-%m-%d')
                    if hasattr(v, 'strftime'):
                        return v.strftime('%Y-%m-%d')
                    return str(v).split()[0]

                def _split_title_body(text: str) -> tuple:
                    text = text.replace('\r\n', '\n').replace('\r', '\n')
                    parts = [p.strip() for p in text.split('\n') if p.strip()]
                    if not parts:
                        return '', ''
                    title = parts[0]
                    body = '\n'.join(parts[1:])
                    return title, body

                nl = _cell(2)
                id_ = _cell(3)
                nl_title, nl_body = _split_title_body(nl)
                id_title, id_body = _split_title_body(id_)

                # image sidecar
                meta = _mededelingen_image_meta()
                key = f"{year}-{idx}"
                rec = meta.get(key)
                image_url = None
                if rec:
                    image_url = f"/mededelingen-image/{year}/{idx}/{rec['filename']}"

                rows.append({
                    'row_index': idx,
                    'category': _cell(0),
                    'type': _cell(1),
                    'nl_title': nl_title,
                    'nl_body': nl_body,
                    'id_title': id_title,
                    'id_body': id_body,
                    'first_date': _date_cell(4),
                    'last_date': _date_cell(5),
                    'event_date': _date_cell(6),
                    'source': _cell(7),
                    'status': _cell(8),
                    'image_url': image_url,
                    'image_name': rec['original_name'] if rec else None,
                })

            return rows

        except Exception as e:
            print(f"Error reading mededelingen rows: {e}")
            return []

    def save_mededelingen_row(self, year: int, row_index: int, data: Dict[str, Any]) -> Dict[str, Any]:
        """Update one row in the year sheet and recompute the Output sheet."""
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        path = MEDEDELINGEN_PATH_TEMPLATE.format(year=year)
        try:
            _, response = self.dbx.files_download(path)
            wb = load_workbook(BytesIO(response.content))

            year_ws = wb[str(year)]
            excel_row = row_index + 2  # header at row 1

            def _set(col, val):
                year_ws.cell(excel_row, col).value = val

            _set(1, data.get('category', ''))
            _set(2, data.get('type', ''))
            _set(3, data.get('nl', ''))
            _set(4, data.get('id', ''))

            # dates written as real dates if possible
            for col_idx, key in [(5, 'first_date'), (6, 'last_date'), (7, 'event_date')]:
                v = data.get(key, '')
                if v:
                    try:
                        d = pd.to_datetime(v)
                        year_ws.cell(excel_row, col_idx).value = d
                    except Exception:
                        year_ws.cell(excel_row, col_idx).value = str(v)
                else:
                    year_ws.cell(excel_row, col_idx).value = None

            _set(8, data.get('source', ''))
            _set(9, data.get('status', ''))

            # Recompute Output sheet from active rows
            output_ws = wb['Output']
            regionale_nl = []
            regionale_id = []
            landelijke_nl = []
            landelijke_id = []

            for r in range(2, year_ws.max_row + 1):
                cat = str(year_ws.cell(r, 1).value or '').strip()
                status = str(year_ws.cell(r, 9).value or '').strip()
                if status.lower() != 'active':
                    continue
                nl = str(year_ws.cell(r, 3).value or '').strip()
                id_ = str(year_ws.cell(r, 4).value or '').strip()
                if cat.lower() == 'regionale':
                    if nl:
                        regionale_nl.append(nl)
                    if id_:
                        regionale_id.append(id_)
                elif cat.lower() == 'landelijke':
                    if nl:
                        landelijke_nl.append(nl)
                    if id_:
                        landelijke_id.append(id_)

            output_ws.cell(2, 2).value = '\n\n'.join(regionale_nl)
            output_ws.cell(2, 3).value = '\n\n'.join(regionale_id)
            output_ws.cell(3, 2).value = '\n\n'.join(landelijke_nl)
            output_ws.cell(3, 3).value = '\n\n'.join(landelijke_id)

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            self.dbx.files_upload(
                out.read(),
                path,
                mode=dropbox.files.WriteMode.overwrite,
                mute=True
            )

            return {'success': True}

        except Exception as e:
            print(f"Error saving mededelingen row: {e}")
            return {'success': False, 'error': str(e)}

    def add_mededelingen_row(self, year: int, data: Dict[str, Any]) -> Dict[str, Any]:
        """Append a new row to the year sheet and recompute Output."""
        from openpyxl import load_workbook

        path = MEDEDELINGEN_PATH_TEMPLATE.format(year=year)
        try:
            _, response = self.dbx.files_download(path)
            wb = load_workbook(BytesIO(response.content))

            year_ws = wb[str(year)]
            new_row = year_ws.max_row + 1

            year_ws.cell(new_row, 1).value = data.get('category', 'Regionale')
            year_ws.cell(new_row, 2).value = data.get('type', 'Others')
            year_ws.cell(new_row, 3).value = data.get('nl', '')
            year_ws.cell(new_row, 4).value = data.get('id', '')

            for col_idx, key in [(5, 'first_date'), (6, 'last_date'), (7, 'event_date')]:
                v = data.get(key, '')
                if v:
                    try:
                        d = pd.to_datetime(v)
                        year_ws.cell(new_row, col_idx).value = d
                    except Exception:
                        year_ws.cell(new_row, col_idx).value = str(v)
                else:
                    year_ws.cell(new_row, col_idx).value = None

            year_ws.cell(new_row, 8).value = data.get('source', 'web')
            year_ws.cell(new_row, 9).value = data.get('status', 'Active')

            # Recompute Output
            output_ws = wb['Output']
            regionale_nl = []
            regionale_id = []
            landelijke_nl = []
            landelijke_id = []

            for r in range(2, year_ws.max_row + 1):
                cat = str(year_ws.cell(r, 1).value or '').strip()
                status = str(year_ws.cell(r, 9).value or '').strip()
                if status.lower() != 'active':
                    continue
                nl = str(year_ws.cell(r, 3).value or '').strip()
                id_ = str(year_ws.cell(r, 4).value or '').strip()
                if cat.lower() == 'regionale':
                    if nl:
                        regionale_nl.append(nl)
                    if id_:
                        regionale_id.append(id_)
                elif cat.lower() == 'landelijke':
                    if nl:
                        landelijke_nl.append(nl)
                    if id_:
                        landelijke_id.append(id_)

            output_ws.cell(2, 2).value = '\n\n'.join(regionale_nl)
            output_ws.cell(2, 3).value = '\n\n'.join(regionale_id)
            output_ws.cell(3, 2).value = '\n\n'.join(landelijke_nl)
            output_ws.cell(3, 3).value = '\n\n'.join(landelijke_id)

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            self.dbx.files_upload(
                out.read(),
                path,
                mode=dropbox.files.WriteMode.overwrite,
                mute=True
            )

            return {'success': True, 'row_index': new_row - 2}

        except Exception as e:
            print(f"Error adding mededelingen row: {e}")
            return {'success': False, 'error': str(e)}

    def get_activiteiten_kalender(self, mededelingen_date: datetime = None) -> List[Dict[str, Any]]:
        """Read Activiteiten Kalender from the year tab (e.g. '2026') of Mededelingen Overzicht.

        Sheet columns (0-indexed, no header row used):
          A(0): Category (Regionale/Landelijke)
          B(1): Type     (Activity/Overlijden/Huwelijks/etc) — used as activity name
          C(2): Details - Nederlands — first line = title, body parsed for time/olv/locatie
          G(6): Event Date (datetime) — used directly for datum
          I(8): Status   (Active / Future) — filter

        Returns list of dicts: datum, tijd, activiteit, olv, locatie  sorted by event date.
        """
        import re

        if mededelingen_date is None:
            year = datetime.now().year
        else:
            year = mededelingen_date.year

        path = MEDEDELINGEN_PATH_TEMPLATE.format(year=year)
        sheet_name = str(year)

        NL_MONTHS = ['jan', 'feb', 'mrt', 'apr', 'mei', 'jun',
                     'jul', 'aug', 'sep', 'okt', 'nov', 'dec']
        NL_MONTH_NAMES = {
            'januari': 'jan', 'februari': 'feb', 'maart': 'mrt', 'april': 'apr',
            'mei': 'mei', 'juni': 'jun', 'juli': 'jul', 'augustus': 'aug',
            'september': 'sep', 'oktober': 'okt', 'november': 'nov', 'december': 'dec',
        }
        MONTHS_PAT = r'(?:' + '|'.join(NL_MONTH_NAMES.keys()) + r')'

        # Multiple dates: "6, 13, 20 en 27 juni" → "6/13/20/27 jun"
        multi_date_re = re.compile(
            rf'((?:\d{{1,2}}(?:[,\s]+(?:en\s+)?)?)+)\s+({MONTHS_PAT})',
            re.IGNORECASE
        )
        # Time range: "van 09.00 tot 11.00 uur" → take first (start) time
        time_range_re = re.compile(r'\bvan\s+(\d{1,2}[:.]\d{2})\s*(?:tot|–|-)', re.IGNORECASE)
        # Single time: "10.30 uur" or "10:30u"
        time_re = re.compile(r'\b(\d{1,2}[:.]\d{2})\s*u(?:ur)?\b', re.IGNORECASE)
        # Free-text time: "na de dienst" / "na de eredienst" / "voor de dienst"
        free_time_re = re.compile(r'\b(na de (?:eredienst|dienst)|voor de (?:eredienst|dienst)|na afloop)\b', re.IGNORECASE)

        olv_re = re.compile(
            r'(?:o\.?l\.?v\.?|onder leiding van|geleid door|gepresenteerd door'
            r'|[Vv]oorganger(?:\s+is)?|waarbij\s+|waarin\s+)'
            r'\s*((?:mw\.|dhr\.|ds\.|br\.|zr\.|dr\.)\s*\S+(?:\s+\S+){0,4}?)'
            r'(?=\s+(?:voorgaat|spreekt|zal\b|op\s+(?:zondag|maandag|\d)|de\s)|[,.\n]|$)',
            re.IGNORECASE
        )
        loc_re = re.compile(
            r'(?:\bin\s+(?:de\s+)?([A-Z][A-Za-zÀ-ÿ\s]+\s+\d+)'
            r'|\bte\s+([A-Z][A-Za-zÀ-ÿ\s]{2,20}?)(?=[,.\n ]|$)'
            r'|\bin\s+(?:de\s+)?([A-Z][A-Za-zÀ-ÿ]{3,20}kerk\b))'
        )

        try:
            print(f"Reading Activiteiten Kalender: {path}, sheet={sheet_name}")
            _, response = self.dbx.files_download(path)
            df = pd.read_excel(BytesIO(response.content), sheet_name=sheet_name, header=None)

            activities = []
            for idx, row in df.iterrows():
                # Filter by Type (col B = index 1): only 'Activity'
                row_type = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
                if row_type.lower() != 'activity':
                    continue

                # col C (index 2): Dutch details text
                content = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ''
                if not content or content.lower() == 'nan':
                    continue

                # col G (index 6): Event Date — already a datetime from Excel
                event_date = row.iloc[6] if len(row) > 6 and pd.notna(row.iloc[6]) else None
                if hasattr(event_date, 'date'):
                    event_date = event_date
                elif isinstance(event_date, str):
                    try:
                        event_date = datetime.strptime(event_date[:10], '%Y-%m-%d')
                    except Exception:
                        event_date = None

                # Filter past events
                if mededelingen_date and event_date:
                    if event_date.date() < mededelingen_date.date():
                        continue

                # Activity title: first line of col C
                first_line = content.split('\n')[0].strip().rstrip(':–-').strip()
                activiteit = first_line

                flat = ' '.join(content.split('\n'))

                # Datum: check for multi-date pattern first ("6, 13, 20 en 27 juni")
                datum = ''
                mdm = multi_date_re.search(flat)
                if mdm:
                    numbers_str = mdm.group(1)
                    month_str = mdm.group(2).lower()
                    month_abbr = NL_MONTH_NAMES.get(month_str, month_str[:3])
                    nums = re.findall(r'\d+', numbers_str)
                    datum = '/'.join(nums) + ' ' + month_abbr
                elif event_date:
                    datum = f"{event_date.day} {NL_MONTHS[event_date.month - 1]}"

                # Tijd: range → start time; free-text → as-is; single time fallback
                tijd = ''
                trm = time_range_re.search(flat)
                if trm:
                    tijd = trm.group(1).replace('.', ':')
                else:
                    ftm = free_time_re.search(flat)
                    if ftm:
                        tijd = ftm.group(1).lower()
                    else:
                        tm = time_re.search(flat)
                        if tm:
                            tijd = tm.group(1).replace('.', ':')

                # o.l.v.
                ovm = olv_re.search(flat)
                olv = ''
                if ovm:
                    raw = ovm.group(1).strip()
                    raw = re.sub(r'\b(ds|br|zr|dr|drs|mr|ir|prof|mw|dhr)\.\s*', r'\1___', raw, flags=re.IGNORECASE)
                    raw = re.sub(r'\b([A-Z])\.\s*', r'\1___', raw)
                    raw = re.split(r',|\.\s+[a-z]|\s+op\s+(?:zaterdag|zondag|maandag|dinsdag|woensdag|donderdag|vrijdag|\d)', raw)[0]
                    olv = raw.replace('___', '. ').strip().rstrip('–- ').strip()

                # Location
                locatie = ''
                flat_lower = flat.lower()
                if re.search(r'\bzoom\b', flat_lower):
                    locatie = 'Zoom'
                elif re.search(r'\blive\b', flat_lower):
                    locatie = 'Live'
                else:
                    for lm in loc_re.finditer(flat):
                        cand = (lm.group(1) or lm.group(2) or lm.group(3) or '').strip().rstrip('.,')
                        if cand and len(cand) <= 40:
                            locatie = cand
                            break
                # Default locatie to Bouwerij 52 when activity is after the service
                if not locatie and re.search(r'na de (?:eredienst|dienst)', tijd, re.IGNORECASE):
                    locatie = 'Bouwerij 52'

                activities.append({
                    'datum': datum,
                    'tijd': tijd,
                    'activiteit': activiteit,
                    'olv': olv,
                    'locatie': locatie,
                    '_sort_date': event_date,
                })

            # Sort by event date ascending
            activities.sort(key=lambda x: x['_sort_date'] or datetime.max)
            for a in activities:
                del a['_sort_date']

            print(f"Activiteiten kalender: {len(activities)} rows loaded from sheet '{sheet_name}'")
            return activities

        except Exception as e:
            print(f"Error reading activiteiten kalender (sheet '{sheet_name}'): {e}")
            return []

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _build_people_map(self, df: pd.DataFrame) -> None:
        """Build a lookup dict: short_name -> {first, last, title}.
        Also builds _predikant_email_map: full_name -> email from rows
        where Short Name is empty but col F (index 5) has a full name.
        """
        self._people_map = {}
        self._predikant_email_map = {}
        for _, row in df.iterrows():
            def _s(col):
                v = row.get(col, '') if isinstance(col, str) else (row.iloc[col] if col < len(row) else '')
                s = str(v).strip() if pd.notna(v) else ''
                return '' if s.lower() == 'nan' else s
            short = _s('Short Name')
            email = _s('Email') or _s(3)
            full  = _s(5)   # col F = full predikant name (may exist even with short name)
            # Index col F full name -> email for predikant lookup regardless of short name
            if full and email:
                self._predikant_email_map[full.lower()] = email
            if not short:
                continue
            self._people_map[short.lower()] = {
                'first_name': _s('First Name'),
                'last_name':  _s('Last Name'),
                'title':      _s('Title'),
                'email':      email,
            }

    def _resolve_name(self, short_name: str) -> str:
        """Convert a short name to 'title First Last' using People tab.
        Tracks unresolved names in _unresolved_names list.
        Lookup is case-insensitive."""
        short_name = short_name.strip()
        if not short_name or not self._people_map:
            if short_name and hasattr(self, '_unresolved_names'):
                self._unresolved_names.append(short_name)
            return short_name
        person = self._people_map.get(short_name.lower())
        if person:
            parts = [person['title'], person['first_name'], person['last_name']]
            return ' '.join(p for p in parts if p and p.lower() != 'nan')
        # Name not found in People tab
        if hasattr(self, '_unresolved_names'):
            self._unresolved_names.append(short_name)
        return short_name

    def _resolve_email(self, short_name: str) -> str:
        """Return email for a short name from People tab.
        Lookup is case-insensitive."""
        short_name = short_name.strip()
        if not short_name or not self._people_map:
            return ''
        person = self._people_map.get(short_name.lower())
        return person.get('email', '') if person else ''

    def _resolve_email_list(self, raw_value: str) -> str:
        """Resolve emails for comma-separated names like 'Bart, Samuel'."""
        if not raw_value or raw_value == '-':
            return ''
        names = [n.strip() for n in raw_value.split(',') if n.strip() and n.strip() != '-']
        emails = []
        for n in names:
            email = self._resolve_email(n)
            if email:
                emails.append(email)
        return ', '.join(emails)

    def _parse_current_sheet(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Parse the CURRENT sheet into a list of service entries."""
        # Find the header row (contains 'DAG' and 'DATUM')
        header_idx = None
        for i in range(min(10, len(df))):
            row_vals = [str(v).strip().upper() for v in df.iloc[i] if pd.notna(v)]
            if 'DAG' in row_vals and 'DATUM' in row_vals:
                header_idx = i
                break

        if header_idx is None:
            print("Could not find header row in CURRENT sheet")
            return []

        # Map column names to indices
        headers = [str(v).strip().upper() if pd.notna(v) else '' for v in df.iloc[header_idx]]
        col = {}
        for idx, h in enumerate(headers):
            h_norm = h.strip().upper().replace('\n', ' ')
            # Canonical mappings
            if h_norm in ('DAG', 'DATUM', 'PREDIKANT', 'OPMERKING', 'TIJD',
                          'BEAMER', 'MUZIEK', 'MULTIMEDIA'):
                col[h_norm] = idx
            elif h_norm in ('OVD', 'OVD.', 'OV D'):
                col['OVD'] = idx
            elif h_norm.startswith('1E') or h_norm in ('1E ONTV', '1E OUDERLING',
                                                        'EERSTE OUDERLING', '1EO'):
                col['1EO'] = idx
        print(f'Takenrooster columns found: {col}')

        entries = []
        for i in range(header_idx + 1, len(df)):
            row = df.iloc[i]
            datum_val = row.iloc[col.get('DATUM', 1)]
            if pd.isna(datum_val):
                continue

            # Parse date
            if isinstance(datum_val, datetime):
                date_obj = datum_val
            else:
                try:
                    date_obj = pd.to_datetime(datum_val)
                except Exception:
                    continue

            def _cell(key, default_col):
                idx2 = col.get(key, default_col)
                v = row.iloc[idx2] if idx2 < len(row) else None
                return str(v).strip() if pd.notna(v) else ''

            dag       = _cell('DAG', 0)
            predikant = _cell('PREDIKANT', 4)
            ovd_short = _cell('OVD', 5)
            opmerking = _cell('OPMERKING', 3)
            # Extract YouTube link from opmerking (e.g. "OLE: https://youtube.com/live/...")
            youtube_link = ''
            if opmerking:
                yt_m = re.search(r'https?://(?:www\.)?(?:youtube\.com/(?:live|watch)|youtu\.be)/[^\s<>"\']*', opmerking)
                if yt_m:
                    youtube_link = yt_m.group(0).rstrip('.,)')
            eo1_short    = _cell('1EO', 6)    # G = index 6 (1e ONTV)
            beamer_short = _cell('BEAMER', 9)  # J = index 9 (BEAMER)
            # Parse TIJD: may be a time object or string like '10:30:00'
            tijd_raw = _cell('TIJD', 2)
            import datetime as _dt
            if hasattr(tijd_raw, 'strftime'):
                tijd = tijd_raw.strftime('%H:%M')
            elif tijd_raw:
                try:
                    t = _dt.time.fromisoformat(str(tijd_raw).split('.')[0])
                    tijd = t.strftime('%H:%M')
                except Exception:
                    tijd = str(tijd_raw)[:5]
            else:
                tijd = '10:30'

            def _resolve_list(key, default_col):
                raw = _cell(key, default_col)
                if not raw or raw == '-':
                    return ''
                names = [n.strip() for n in raw.split(',') if n.strip() and n.strip() != '-']
                resolved = []
                for n in names:
                    r = self._resolve_name(n)
                    resolved.append(r or n)
                return ', '.join(resolved)

            def _resolve_name_list(raw_value):
                """Resolve a comma-separated list of names like _resolve_list but returns individual unresolved tracking."""
                if not raw_value or raw_value == '-':
                    return ''
                names = [n.strip() for n in raw_value.split(',') if n.strip() and n.strip() != '-']
                resolved = []
                for n in names:
                    r = self._resolve_name(n)
                    resolved.append(r or n)
                return ', '.join(resolved)

            # Track unresolved names for this entry
            self._unresolved_names = []

            muziek      = _resolve_list('MUZIEK', 10)
            voorzangers = _resolve_list('VOORZANGERS', 11)
            multimedia  = _resolve_list('MULTIMEDIA', 12)
            knd_raw     = _cell('KND', 7)
            tieners_raw = _cell('TIENERS', 8)
            knd         = _resolve_name_list(knd_raw) if knd_raw and knd_raw != '-' else ''
            tieners     = _resolve_name_list(tieners_raw) if tieners_raw and tieners_raw != '-' else ''

            # Resolve OVD/1EO/Beamer names (tracks unresolved)
            ovd_full    = self._resolve_name(ovd_short) or ovd_short
            eo1_full    = self._resolve_name(eo1_short) or eo1_short
            # BEAMER may contain comma-separated names like "Bart, Samuel"
            beamer_full = _resolve_name_list(beamer_short) if beamer_short and beamer_short != '-' else ''

            # Get unique unresolved names
            unresolved = list(dict.fromkeys(self._unresolved_names))  # preserve order, remove duplicates

            # Normalize predikant salutation to lowercase (ds./zr./br./mw./mevr./dhr./dr.)
            predikant = re.sub(
                r'^(Ds|Zr|Br|Mw|Mevr|Dhr|Dr)\.',
                lambda m: m.group(0).lower(),
                predikant,
                flags=re.IGNORECASE
            )

            # Resolve predikant email from People tab
            pred_email  = self._predikant_email_map.get(predikant.lower(), '')
            ovd_email   = self._resolve_email(ovd_short)
            eo1_email   = self._resolve_email(eo1_short)
            beamer_email = self._resolve_email_list(beamer_short)

            entries.append({
                'date':            date_obj,
                'dag':             dag,
                'predikant':       predikant,
                'predikant_email': pred_email,
                'ovd':             ovd_full,
                'ovd_email':       ovd_email,
                '1eo':             eo1_full,
                '1eo_email':       eo1_email,
                'beamer':          beamer_full,
                'beamer_email':    beamer_email,
                'opmerking':       opmerking,
                'youtube_link':    youtube_link,
                'tijd':            tijd,
                'muziek':          muziek,
                'voorzangers':     voorzangers,
                'multimedia':      multimedia,
                'knd':             knd,
                'tieners':         tieners,
                'unresolved_names': unresolved,
            })

        return entries
